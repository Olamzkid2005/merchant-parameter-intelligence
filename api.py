"""
api.py — FastAPI backend for the Merchant Intelligence React frontend.

Exposes the merchant_intelligence search engine over HTTP so the React
app (web/) can query the SQLite registry without touching the DB directly.

Endpoints
---------
GET  /api/health          — liveness probe
GET  /api/stats           — total record count
POST /api/search          — {query, limit} → scored results
POST /api/batch           — {merchants: [...]} → best-match rows
POST /api/batch/export    — same as /api/batch but returns an .xlsx file
"""
import json
import logging
import re
import sys
import time
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


def _audit(action: str, scope: Optional[str] = None,
           detail: Optional[str] = None) -> None:
    """Best-effort audit-trail append (roadmap #1) — never breaks the
    request; audit.record() already swallows its own failures. Actor comes
    from the auth middleware context (the session username) or "local"."""
    try:
        from merchant_intelligence import audit
        from merchant_intelligence import auth
        audit.record(action, scope=scope, detail=detail,
                     actor=auth.current_actor())
    except Exception:  # noqa: BLE001
        pass

PROJECT_ROOT = Path(__file__).resolve().parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
# CLI tools (report.py, data_quality.py, reconcile.py) live in scripts/
_SCRIPTS_DIR = PROJECT_ROOT / "scripts"
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

import pandas as pd
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response
from pydantic import BaseModel, Field

from merchant_intelligence import MerchantSearch, config
from merchant_intelligence.entity import EntityResolver
from merchant_intelligence.fuzzy import (damerau_levenshtein_similarity,
                                         token_sort_ratio)

app = FastAPI(title="Merchant Intelligence API", version="1.0.0")

# Allow the Vite dev server to call the API directly (a Vite proxy is also
# configured, but CORS keeps the door open for other origins).
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def security_middleware(request: Request, call_next):
    """Opt-in authN/Z + field masking (roadmap #1, slice 2). When access
    control is disabled (the default) every request passes through
    untouched — the desktop-tool experience is unchanged."""
    from merchant_intelligence import auth
    if not auth.enabled():
        return await call_next(request)
    path = request.url.path
    if path in auth.EXEMPT_PATHS:
        return await call_next(request)
    token = request.cookies.get("mi_session")
    session = auth.get_session(token) if token else None
    if not session:
        return JSONResponse(status_code=401,
                            content={"detail": "authentication required"})
    if not auth.require(path, request.method, session["role"]):
        return JSONResponse(status_code=403,
                            content={"detail": "insufficient role for this action"})
    auth._current_actor.set(session["username"])
    resp = await call_next(request)
    # Field-level masking at the API boundary for viewer sessions.
    if session["role"] == "viewer" and "application/json" in (
            resp.headers.get("content-type") or ""):
        body = b"".join([chunk async for chunk in resp.body_iterator])
        # Drop length/encoding headers — the masked body has a different
        # size; letting Response recompute content-length avoids truncated
        # reads (IncompleteRead) on the client.
        headers = {k: v for k, v in resp.headers.items()
                   if k.lower() not in ("content-length", "content-encoding")}
        try:
            masked = json.dumps(auth.mask_payload(json.loads(body)),
                                default=str).encode("utf-8")
            return Response(content=masked, status_code=resp.status_code,
                            headers=headers, media_type="application/json")
        except Exception:  # noqa: BLE001 — never break a viewer response
            return Response(content=body, status_code=resp.status_code,
                            headers=headers, media_type=resp.media_type)
    return resp


_searcher = None
_resolver = None
_profiler = None


def _multi_identifier_query(query: str) -> Optional[List[str]]:
    """Return the identifier values when a query pastes 2+ of them at once.

    ``"2ISW2587 2ISW2586"`` (two TIDs) or ``"MX183544 MX183545"`` is a
    batch-resolve request: each identifier must be searched on its own and
    the mentions merged — treating the whole string as one search blob makes
    FTS fuzzy-match unrelated rows (e.g. UBTH records matching on loose
    token overlap). A single identifier or a plain name query returns None
    and searches normally.
    """
    try:
        from merchant_intelligence.tasks.parser import parse_identifiers
        ident = parse_identifiers(query)
        values = [v for vals in ident.values() for v in vals]
        values = list(dict.fromkeys(values))  # dedupe, preserve order
        if len(values) >= 2:
            return values
    except Exception:
        pass
    return None


def _search_with_multi(query: str, fetch: int, min_score: float) -> list:
    """Search, OR-ing the identifiers when the query pastes 2+ of them.

    Every returned SearchResult keeps its own match metadata (matched_field /
    matched_value) from its identifier's search, and rows are deduped by
    record id so a value stored in two columns never double-counts.
    """
    searcher = get_searcher()
    multi = _multi_identifier_query(query)
    if not multi:
        return searcher.search(query, limit=fetch, min_score=min_score)
    merged: Dict[Any, Any] = {}
    for v in multi:
        # Each identifier is searched on its own; only exact/high-confidence
        # hits count so the loose fuzzy tier (UBTH/terminal-serial noise that
        # matches a bare TID's digits) never pollutes a batch-resolve result.
        for r in searcher.search(v, limit=min(fetch, 25), min_score=min_score):
            if r.match_type not in ("Exact Match", "High Confidence"):
                continue
            rid = r.record.get("id")
            if rid is not None and rid not in merged:
                merged[rid] = r
    return list(merged.values())


def get_searcher() -> MerchantSearch:


    global _searcher
    if _searcher is None:
        _searcher = MerchantSearch()
    return _searcher


def get_resolver() -> EntityResolver:
    global _resolver
    if _resolver is None:
        _resolver = EntityResolver()
    return _resolver


def get_profiler():
    """Return the cached MerchantProfile singleton (avoids re-opening the
    DatabaseManager/EntityResolver/Matcher connections per request)."""
    global _profiler
    if _profiler is None:
        from merchant_intelligence.profile import MerchantProfile
        _profiler = MerchantProfile()
    return _profiler


class SearchRequest(BaseModel):
    query: str
    limit: int = Field(20, ge=1, le=100)
    min_score: float = Field(0.0, ge=0.0, le=100.0)
    offset: int = Field(0, ge=0)


class BatchRequest(BaseModel):
    merchants: list[str] = Field(default_factory=list)


class LearnRequest(BaseModel):
    query: str
    merchant_name: str


class EntityRequest(BaseModel):
    query: str
    depth: int = Field(2, ge=1, le=3)
    max_nodes: int = Field(120, ge=5, le=300)


class QuickMatchRequest(BaseModel):
    identifiers: list[str] = Field(default_factory=list)


class ProfileRequest(BaseModel):
    query: str
    max_members: int = Field(200, ge=1, le=1000)


class CompareRequest(BaseModel):
    query_a: str
    query_b: str
    max_members: int = Field(200, ge=1, le=1000)


@app.get("/api/health")
def health():
    return {"status": "ok"}


class LoginRequest(BaseModel):
    username: str = ""
    password: str = ""


class AuthConfigRequest(BaseModel):
    enabled: Optional[bool] = None
    session_ttl_hours: Optional[float] = None


class AuthUserRequest(BaseModel):
    username: str = ""
    password: str = ""
    role: str = "viewer"


class AuthPasswordRequest(BaseModel):
    username: str = ""
    password: str = ""


@app.post("/api/auth/login")
def auth_login(req: LoginRequest):
    """Username + password -> expiring session cookie (mi_session)."""
    from merchant_intelligence import auth
    username = req.username.strip()
    if not username or not req.password:
        raise HTTPException(status_code=400,
                            detail="username and password are required")
    cfg = auth.load_config()
    user = next((u for u in cfg["users"]
                 if u["username"].lower() == username.lower()), None)
    if not user or not auth.verify_password(req.password, user["salt"],
                                            user["hash"]):
        raise HTTPException(status_code=401,
                            detail="invalid username or password")
    token = auth.create_session(user["username"], user["role"])
    ttl = float(cfg.get("session_ttl_hours", 12)) * 3600
    resp = JSONResponse({"ok": True, "user": user["username"],
                         "role": user["role"]})
    resp.set_cookie("mi_session", token, httponly=True, samesite="lax",
                    path="/", max_age=int(ttl))
    return resp


@app.post("/api/auth/logout")
def auth_logout(request: Request):
    """Destroy the current session and clear the cookie."""
    from merchant_intelligence import auth
    token = request.cookies.get("mi_session")
    if token:
        auth.destroy_session(token)
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("mi_session", path="/")
    return resp


@app.get("/api/auth/me")
def auth_me(request: Request):
    """Auth status for the UI: enabled + who am I (if logged in)."""
    from merchant_intelligence import auth
    cfg = auth.load_config()
    if not cfg.get("enabled"):
        return {"enabled": False, "authenticated": False}
    token = request.cookies.get("mi_session")
    session = auth.get_session(token) if token else None
    if not session:
        return {"enabled": True, "authenticated": False}
    return {"enabled": True, "authenticated": True,
            "user": session["username"], "role": session["role"]}


@app.get("/api/auth/config")
def auth_config():
    """Security config for the Rule Engine card (users shown without
    hashes). Write endpoints are admin-gated once access control is on."""
    from merchant_intelligence import auth
    cfg = auth.load_config()
    return {"enabled": bool(cfg.get("enabled")),
            "session_ttl_hours": cfg.get("session_ttl_hours", 12),
            "users": [{"username": u["username"], "role": u["role"]}
                      for u in cfg["users"]]}


@app.post("/api/auth/config")
def auth_save_config(req: AuthConfigRequest):
    """Toggle access control and/or the session TTL. Enabling with zero
    users would lock everyone out, so that is refused."""
    from merchant_intelligence import auth
    cfg = auth.load_config()
    if req.enabled is not None:
        cfg["enabled"] = bool(req.enabled)
        if cfg["enabled"] and not cfg["users"]:
            raise HTTPException(
                status_code=400,
                detail="add a user before enabling access control")
    if req.session_ttl_hours is not None:
        cfg["session_ttl_hours"] = max(1, min(168,
                                               float(req.session_ttl_hours)))
    auth.save_config(cfg)
    return {"ok": True, "enabled": cfg["enabled"],
            "session_ttl_hours": cfg["session_ttl_hours"]}


@app.post("/api/auth/users")
def auth_add_user(req: AuthUserRequest):
    """Create a user (bootstrap path works while access control is off)."""
    from merchant_intelligence import auth
    username = req.username.strip()
    if not username or not req.password:
        raise HTTPException(status_code=400,
                            detail="username and password are required")
    if req.role not in auth.ROLES:
        raise HTTPException(status_code=400,
                            detail=f"role must be one of {auth.ROLES}")
    if len(req.password) < 8:
        raise HTTPException(status_code=400,
                            detail="password must be at least 8 characters")
    cfg = auth.load_config()
    if any(u["username"].lower() == username.lower()
           for u in cfg["users"]):
        raise HTTPException(status_code=400,
                            detail="username already exists")
    salt, pw_hash = auth.hash_password(req.password)
    cfg["users"].append({"username": username, "role": req.role,
                          "salt": salt, "hash": pw_hash})
    auth.save_config(cfg)
    return {"ok": True,
            "users": [{"username": u["username"], "role": u["role"]}
                      for u in cfg["users"]]}


@app.delete("/api/auth/users")
def auth_remove_user(req: AuthUserRequest):
    """Remove a user; disabling access control when the last user goes."""
    from merchant_intelligence import auth
    username = req.username.strip()
    cfg = auth.load_config()
    before = len(cfg["users"])
    cfg["users"] = [u for u in cfg["users"]
                     if u["username"].lower() != username.lower()]
    if len(cfg["users"]) == before:
        raise HTTPException(status_code=404, detail="username not found")
    if not cfg["users"]:
        cfg["enabled"] = False  # never leave a locked box
    auth.save_config(cfg)
    return {"ok": True, "enabled": cfg["enabled"],
            "users": [{"username": u["username"], "role": u["role"]}
                      for u in cfg["users"]]}


@app.put("/api/auth/password")
def auth_reset_password(req: AuthPasswordRequest):
    """Reset a user's password (admin-gated when access control is on)."""
    from merchant_intelligence import auth
    username = req.username.strip()
    if not username or not req.password:
        raise HTTPException(status_code=400,
                            detail="username and password are required")
    if len(req.password) < 8:
        raise HTTPException(status_code=400,
                            detail="password must be at least 8 characters")
    cfg = auth.load_config()
    user = next((u for u in cfg["users"]
                 if u["username"].lower() == username.lower()), None)
    if not user:
        raise HTTPException(status_code=404, detail="username not found")
    salt, pw_hash = auth.hash_password(req.password)
    user["salt"], user["hash"] = salt, pw_hash
    auth.save_config(cfg)
    return {"ok": True}


@app.post("/api/profile")
def profile(req: ProfileRequest):
    """Merchant 360° profile — everything the registry knows about a fragment.

    Accepts ANY fragment (name, email, phone, MX code, TID, account number…)
    and returns the seed match plus an aggregation across the whole entity
    family: every unique email/phone/TID/MX/address/contact, the distinct
    merchant-name variants, the sources, and the full linked rows.
    """
    query = req.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="query is required")
    _audit("profile", json.dumps({"query": query}))
    return get_profiler().build(query, max_members=req.max_members)


class TimelineRequest(BaseModel):
    query: str


@app.post("/api/timeline")
def timeline(req: TimelineRequest):
    """Per-terminal timeline for a merchant fragment (build-time events).

    Accepts any fragment (name, phone, MX code, TID, account number) and
    resolves it to the terminal key(s) the registry actually stores, then
    returns every derived event from the merchant_events table:
      - first_seen / last_seen  (onboarding + most recent trace)
      - name_variant            every distinct name, with its source file
      - account_change          old->new transitions from the Change sheet

    Resolution is DB-grounded: keys only exist when the registry stores them.
    """
    import sqlite3
    from merchant_intelligence.enrich import keys_for_query, timeline_for
    query = req.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="query is required")
    conn = sqlite3.connect(str(config.active_db()))
    conn.row_factory = sqlite3.Row
    try:
        keys = keys_for_query(conn, query, limit=8)
        out = []
        for field, key in keys:
            events = timeline_for(conn, key)
            if events:
                out.append({
                    "key_field": field,
                    "terminal_key": key,
                    "event_count": len(events),
                    "events": events,
                })
        return {"query": query, "resolved": len(out), "terminals": out}
    finally:
        conn.close()


@app.post("/api/compare")
def compare(req: CompareRequest):
    """Compare two merchants' 360° profiles side by side.

    Builds both profiles (any fragments work — name, email, phone, MX code,
    TID, account number…) and diffs them: which identifiers the two share,
    a per-field match/overlap/differ table, and how many family rows overlap.
    """
    query_a = req.query_a.strip()
    query_b = req.query_b.strip()
    if not query_a or not query_b:
        raise HTTPException(status_code=400, detail="query_a and query_b are required")
    return get_profiler().compare(query_a, query_b, max_members=req.max_members)


@app.get("/api/stats")
def stats():
    import sqlite3
    try:
        conn = sqlite3.connect(str(config.active_db()))
        total = conn.execute("SELECT COUNT(*) FROM merchants").fetchone()[0]
        conn.close()
    except Exception:
        total = 0
    return {"total_records": total}


@app.post("/api/search")
def search(req: SearchRequest):
    query = req.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="query is required")
    _audit("search", json.dumps({"query": query,
                                 "offset": req.offset, "limit": req.limit}))
    t0 = time.perf_counter()
    fetch = min(req.offset + req.limit, 200)
    results = _search_with_multi(query, fetch, req.min_score)
    total = len(results)
    page = results[req.offset:req.offset + req.limit]
    elapsed_ms = (time.perf_counter() - t0) * 1000
    # Remember this search as follow-up context — a subsequent "get the tids
    # for the above merchant" resolves against it (DB-grounded: the top
    # result's own identifiers + canonical name, never the raw query text).
    try:
        if results:
            top = results[0].to_dict()
            ids = {}
            for k in ("tid", "mxcode", "phone", "email"):
                v = top.get(k)
                if v:
                    ids.setdefault(k, []).append(str(v))
            name = top.get("merchant_name")
            from merchant_intelligence import tasks as _tasks
            _tasks.remember_entities(identifiers=ids,
                                     names=[name] if name else None)
    except Exception:
        pass
    # Feed the search into the self-improvement loop (rephrase detection +
    # outcome stats). A search with zero results is a candidate "failed
    # phrasing" — if the same merchant is re-asked shortly after, the search
    # is tagged rephrased and the follow-up's intent becomes the correction.
    try:
        from merchant_intelligence import feedback as _fb
        _fb.log_request(kind="search", text=query, intent="", rows=total,
                        entity_sig=[query.upper().strip()])
    except Exception:
        pass
    return {
        "query": query,
        "count": len(page),
        "total": total,
        "offset": req.offset,
        "elapsed_ms": round(elapsed_ms, 1),
        "results": [r.to_dict() for r in page],
    }


def _run_batch(merchants: list[str]) -> list[dict]:
    """Run one search per merchant, returning the best match per input."""
    searcher = get_searcher()
    rows = []
    for m in merchants:
        res = searcher.search(m, limit=1, min_score=0)
        best = res[0] if res else None
        rec = best.record if best else {}
        rows.append({
            "input": m,
            "best_match": rec.get("merchant_name", ""),
            "score": round(best.overall_score / 10, 1) if best else 0,
            "match_type": best.match_type if best else "Not Found",
            "email": rec.get("email", ""),
            "phone": rec.get("phone", ""),
            "tid": rec.get("tid", ""),
            "mxcode": rec.get("mxcode", ""),
            "sheet": rec.get("sheet_name", ""),
            # Key-merchant roots for the matched name so Batch rows carry the
            # same family badge as the Search page (clickable -> profile).
            "key_merchants": _key_merchants_for(rec.get("merchant_name", "")),
        })
    return rows


@app.post("/api/entity")
def entity(req: EntityRequest):
    """Entity graph: merchant family + BFS relationship graph.

    Returns the family (records linked to the seed by shared identifiers),
    the raw graph (nodes + edges for visualisation) and alias candidates
    that the frontend can offer to teach the engine.
    """
    query = req.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="query is required")
    t0 = time.perf_counter()
    resolver = get_resolver()
    family = resolver.family_of(query)
    graph = resolver.graph(query, depth=req.depth, max_nodes=req.max_nodes)
    elapsed_ms = (time.perf_counter() - t0) * 1000

    # Annotate each family member with a name-similarity % vs the seed so the
    # frontend can rank node cards (mirrors the Streamlit tab behaviour).
    members = []
    for m in family.get("members", []):
        mname = str(m.get("merchant_name", ""))
        sim = token_sort_ratio(query, mname)
        m = dict(m)
        m["match_pct"] = max(int(sim * 100), 30) if mname else 0
        members.append(m)
    family["members"] = members

    return {
        "seed": query,
        "elapsed_ms": round(elapsed_ms, 1),
        "family": family,
        "graph": graph,
    }


@app.post("/api/search/export")
def search_export(req: SearchRequest):
    """Export the current search view as an Excel workbook."""
    query = req.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="query is required")
    _audit("export", json.dumps({"kind": "search", "query": query}))
    fetch = min(req.offset + req.limit, 200)
    results = _search_with_multi(query, fetch, req.min_score)
    page = results[req.offset:req.offset + req.limit]
    rows = []
    for r in page:
        rec = r.record
        rows.append({
            "Score": round(r.overall_score / 10, 1),
            "Match Type": r.match_type,
            "Merchant Name": rec.get("merchant_name", ""),
            "TID": rec.get("tid", ""),
            "MX Code": rec.get("mxcode", ""),
            "Email": rec.get("email", ""),
            "Phone": rec.get("phone", ""),
            "Contact Name": rec.get("contact_name", ""),
            "Account Name": rec.get("account_name", ""),
            "Onboarded": rec.get("onboarded_date", ""),
            "Sheet": rec.get("sheet_name", ""),
            "Matched By": r.identifier_hit or "",
            # Build-time data-quality score (0-100) so weak records are
            # visible in the exported workbook too.
            "Quality": rec.get("quality_score", 100),
        })
    df = pd.DataFrame(rows)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Search Results", index=False)
        _style_workbook(writer.book)
    return Response(
        buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="search_results.xlsx"'},
    )


@app.get("/api/idclass/debug")
def idclass_debug(values: str = "", text: str = "", limit: int = 50):
    """Debug the DB-rooted identifier classifier (idclass.py).

    GET /api/idclass/debug?values=MX184380,2103O338,5180857349
    GET /api/idclass/debug?text=MX184380 2103O338 FELIX

    For each pasted token, shows WHY it classified the way it did:
      source         db_membership | shape_rule | rejected | unknown
      in_db_columns  the actual registry columns storing that value
      shape_rule     the shape pattern that matched (non-DB values)
      reason         human-readable explanation
    Plus diagnostics about the index itself (DB path, distinct tokens,
    per-kind counts) so you can verify which DB the engine is reading.
    """
    from merchant_intelligence.idclass import get_classifier

    tokens = []
    if values:
        tokens += [v.strip() for v in values.split(",") if v.strip()]
    if text:
        tokens += [t for t in re.split(r"[\s,;]+", text) if t.strip()]
    # Dedup preserving order, cap the result set.
    seen, uniq = set(), []
    for t in tokens:
        key = t.upper()
        if key not in seen:
            seen.add(key)
            uniq.append(t)
    uniq = uniq[:max(1, min(limit, 200))]

    classifier = get_classifier()
    stats = classifier.index_stats()
    return {
        "index": stats,
        "count": len(uniq),
        "results": [classifier.inspect(t) for t in uniq],
    }


def _prefix_edit_similarity(typed: str, candidate: str) -> float:
    """0..1 similarity of the TYPED prefix against the START of a candidate.

    Unlike token_set_ratio (which splits on spaces and can't compare a
    single unbroken token), this measures how many edits turn the typed
    prefix into a PREFIX of the candidate bucket key:
      'MEDPLUZ' -> 'MEDPLUS'  1 edit (substitution)  -> high
      'KONGOPAY' -> 'KONGAPAY' 1 edit                 -> high
      'lagoon watr' -> 'LAGOON WATERS' 1 edit         -> high
      'ZZZZNOTREAL' -> any key  ~9 edits             -> ~0

    Transpositions count as ONE edit (Damerau-Levenshtein), so 'MEDPLUS' vs
    'MEDLPUS' also recovers. Returns 0 when the candidate doesn't start with
    a near-edit version of the typed prefix (no accidental mid-name hits).
    """
    typed = (typed or "").upper().strip()
    candidate = (candidate or "").upper().strip()
    if not typed or not candidate or len(typed) < 3:
        return 0.0
    # DELIBERATE TRADEOFF: candidates must share the typed FIRST character.
    # Real typos almost always hit mid-word (MEDPLUZ), and this gate prunes
    # ~96% of the 13k-key scan before any distance call — keeping the
    # keystroke path sub-millisecond. A first-letter typo ('edplus') does
    # NOT recover here; the full search still finds the merchant, and the
    # autocomplete dropdown just stays silent for that rare case.
    if candidate[0] != typed[0]:
        return 0.0
    limit = min(len(typed), len(candidate))
    sim = damerau_levenshtein_similarity(typed[:limit], candidate[:limit])
    if sim < 0.72:
        return 0.0
    return sim


def _fuzzy_prefix_suggestions(prefix: str, keys, limit: int) -> list:
    """Character-level fuzzy tier: recover bucket keys whose START is a
    typo'd version of the typed prefix (single-token typos the token-based
    tier cannot see). Returns candidates sorted best-first, best first."""
    scored = []
    for k in keys:
        sim = _prefix_edit_similarity(prefix, k)
        if sim > 0.0:
            scored.append((sim, k))
    scored.sort(key=lambda t: (-t[0], len(t[1])))
    return [k for _s, k in scored[:limit]]


@app.get("/api/autocomplete")
def autocomplete(prefix: str = "", limit: int = 8):
    """Live typeahead suggestions from the normalized name_buckets table.

    GET /api/autocomplete?prefix=lagoon&limit=8

    Returns merchant bucket keys whose canonical form starts with the
    prefix ("lagoon wat" -> ["LAGOON WATERS", ...]). Backed by one indexed
    LIKE query on the bucket_key primary key, so it is fast enough to fire
    on every keystroke. Ensures the bucket table is built lazily on first
    use so existing databases work without a rebuild.

    When the exact-prefix tier comes up short, two fuzzy fallbacks recover
    near-misses (both are niceties — never break the page):
      1. token_set_ratio scan over the distinct-name index — recovers
         multi-word phrases with added/missing tokens ("lagoon water ent").
      2. Damerau-Levenshtein PREFIX scan — recovers single-token typos
         ("medpluz" -> MEDPLUS, "kongopay" -> KONGAPAY) that the token
         splitter cannot see.
    """
    prefix = (prefix or "").strip()
    if not prefix:
        return {"prefix": prefix, "suggestions": []}
    try:
        db = get_searcher().matcher.db
        db.ensure_buckets()
        suggestions = db.autocomplete(prefix, limit=max(1, min(limit, 15)))
        if len(suggestions) < limit:
            keys = db.bucket_keys()
            if keys:
                from rapidfuzz import fuzz as _rf_fuzz
                from rapidfuzz import process as _rf_process
                # 1) token_set_ratio (NOT WRatio — its partial component
                #    over-matches garbage like 'ZZZZNOTREAL' against short keys)
                fuzzy = [k for k, _s, _i in _rf_process.extract(
                    prefix, keys, scorer=_rf_fuzz.token_set_ratio,
                    limit=max(1, min(limit, 15)), score_cutoff=70)
                    if k not in suggestions]
                suggestions = (suggestions + fuzzy)[:max(1, min(limit, 15))]
        # 2) character-level prefix tier: single-token typos the token
        #    splitter cannot see ('medpluz' -> MEDPLUS). Always runs so a
        #    typo'd first keypress recovers even when the token tier found
        #    nothing.
        if len(suggestions) < limit:
            keys = db.bucket_keys()
            if keys:
                extra = [k for k in _fuzzy_prefix_suggestions(prefix, keys, limit)
                         if k not in suggestions]
                suggestions = (suggestions + extra)[:max(1, min(limit, 15))]
    except Exception:
        suggestions = []  # autocomplete is a nicety — never break the page
    return {"prefix": prefix, "suggestions": suggestions}


@app.post("/api/suggest")
def suggest(req: SearchRequest):
    """Did-you-mean suggestions when a query yields no confident hits.

    Generates variants of the query (dropped tokens, common typo fixes,
    generic-word removal) and returns the ones that DO find records.
    """
    query = req.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="query is required")
    searcher = get_searcher()
    base = searcher.search(query, limit=3, min_score=config.POSSIBLE_THRESHOLD)
    if base:
        return {"query": query, "suggestions": []}

    suggestions = []
    seen = set()

    def try_add(q):
        q = (q or "").strip()
        if not q or q.lower() in seen:
            return
        seen.add(q.lower())
        res = searcher.search(q, limit=1, min_score=config.POSSIBLE_THRESHOLD)
        if res:
            suggestions.append({
                "query": q,
                "score": round(res[0].overall_score / 10, 1),
                "best_match": res[0].record.get("merchant_name", ""),
            })

    tokens = query.split()
    # 1. Drop each token in turn
    for i in range(len(tokens)):
        try_add(" ".join(tokens[:i] + tokens[i + 1:]))
    # 2. Common spelling fixes (canonicalize applies these at token level
    #    already; here we also retry the RAW query with each correction in
    #    case the fix changes which DB rows are retrievable).
    fixes = dict(config.TYPO_FIXES)
    fixes["BIDGBENGA"] = "BIDDEL"  # query-side name, not a typo of a word
    ql = query.lower()
    for bad, good in fixes.items():
        if bad.lower() in ql:
            try_add(ql.replace(bad.lower(), good.lower()))
    # 3. Strip generic words entirely
    sig = [t for t in tokens if t.upper() not in config.GENERIC_WORDS]
    if len(sig) != len(tokens) and sig:
        try_add(" ".join(sig))

    return {"query": query, "suggestions": suggestions[:6]}


@app.post("/api/similar")
def similar(req: SearchRequest):
    """Similar merchants: records linked to the query by shared identifiers
    (entity family) — different names that resolve to the same merchant."""
    query = req.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="query is required")
    resolver = get_resolver()
    family = resolver.family_of(query)
    members = family.get("members", [])
    seen = set()
    out = []
    for m in members:
        name = m.get("merchant_name", "") or ""
        key = name.upper()
        if not name or key in seen:
            continue
        seen.add(key)
        out.append({
            "merchant_name": name,
            "sheet": m.get("sheet_name", ""),
            "email": m.get("email", ""),
            "phone": m.get("phone", ""),
            "mxcode": m.get("mxcode", ""),
            "tid": m.get("tid", ""),
            "link_reasons": m.get("link_reasons", []),
            # Key-merchant roots (same engine the search badge uses) so the
            # Similar/Related panel can show the same family chip + click.
            "key_merchants": _key_merchants_for(name),
        })
    return {"query": query, "count": len(out), "similar": out[:req.limit]}


@app.get("/api/duplicates")
def duplicates(limit: int = 100):
    """Duplicate merchant clusters: the same merchant name appearing on
    multiple DB rows (across sheets/files), grouped with locations."""
    import sqlite3
    conn = sqlite3.connect(str(config.active_db()))
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    groups = c.execute(
        "SELECT UPPER(merchant_name) name_key, COUNT(*) n "
        "FROM merchants WHERE merchant_name != '' "
        "GROUP BY name_key HAVING n > 1 ORDER BY n DESC LIMIT ?",
        (limit,),
    ).fetchall()
    clusters = []
    for g in groups:
        rows = c.execute(
            "SELECT id, merchant_name, sheet_name, row_number, tid, mxcode, email "
            "FROM merchants WHERE UPPER(merchant_name) = ?",
            (g["name_key"],),
        ).fetchall()
        clusters.append({
            "merchant_name": rows[0]["merchant_name"],
            "occurrences": len(rows),
            "sheets": sorted({r["sheet_name"] for r in rows}),
            "records": [{
                "id": r["id"], "sheet": r["sheet_name"],
                "row": r["row_number"], "tid": r["tid"],
                "mxcode": r["mxcode"], "email": r["email"],
            } for r in rows],
        })
    conn.close()
    return {"count": len(clusters), "clusters": clusters}


@app.get("/api/aliases")
def aliases():
    """Alias review queue: manual + learned aliases with approval status."""
    engine = get_searcher().matcher.alias_engine
    learned = engine.review_items()
    manual = engine.manual_items()
    return {
        "learned": learned,
        "manual": manual,
        "counts": {
            "learned": len(learned),
            "manual": len(manual),
            "pending": sum(1 for i in learned if i["status"] == "pending"),
            "approved": sum(1 for i in learned if i["status"] == "approved"),
        },
    }


class AliasAction(BaseModel):
    alias: str
    canonical: str


@app.post("/api/aliases/approve")
def alias_approve(req: AliasAction):
    """Approve a learned alias in the review queue."""
    engine = get_searcher().matcher.alias_engine
    return {"ok": engine.approve(req.alias, req.canonical)}


@app.post("/api/aliases/reject")
def alias_reject(req: AliasAction):
    """Reject (forget) a learned alias in the review queue."""
    engine = get_searcher().matcher.alias_engine
    return {"ok": engine.forget(req.alias, req.canonical)}


@app.post("/api/report")
def report(req: BatchRequest):
    """Phase 9 Merchant Intelligence Report — full multi-sheet preview."""
    from report import build_report
    merchants = [m.strip() for m in req.merchants if m.strip()][:1000]
    if not merchants:
        raise HTTPException(status_code=400, detail="merchants is empty")
    t0 = time.perf_counter()
    rep = build_report(merchants, top_n=3)
    elapsed = time.perf_counter() - t0

    def recs(key):
        df = rep[key]
        return df.to_dict(orient="records") if df is not None and not df.empty else []

    keys = ["exact", "high", "possible", "emails", "phones",
            "contacts", "addresses", "duplicates", "not_found"]
    return {
        "count": len(merchants),
        "elapsed_s": round(elapsed, 2),
        "summary": recs("summary"),
        "sheet_counts": {k: len(recs(k)) for k in keys},
        "sheets": {k: recs(k) for k in keys},
    }


@app.post("/api/report/export")
def report_export(req: BatchRequest):
    """Build the Phase 9 report and download it as a multi-sheet workbook."""
    from report import build_report, SHEET_ORDER, SHEET_NAMES
    merchants = [m.strip() for m in req.merchants if m.strip()][:1000]
    if not merchants:
        raise HTTPException(status_code=400, detail="merchants is empty")
    _audit("export", json.dumps({"kind": "report", "count": len(merchants)}))
    rep = build_report(merchants, top_n=3)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for key in SHEET_ORDER:
            df = rep[key]
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=SHEET_NAMES[key], index=False)
            else:
                pd.DataFrame({"Note": ["No records"]}).to_excel(
                    writer, sheet_name=SHEET_NAMES[key], index=False)
        _style_workbook(writer.book)
    return Response(
        buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Merchant_Intelligence_Report.xlsx"'},
    )


@app.post("/api/learn")
def learn(req: LearnRequest):
    """Teach the alias engine a new query -> merchant mapping (Phase 10)."""
    query = req.query.strip()
    merchant = req.merchant_name.strip()
    if not query or not merchant:
        raise HTTPException(status_code=400, detail="query and merchant_name are required")
    # Guard: only persist mappings to merchants that exist in the registry.
    import sqlite3
    conn = None
    try:
        conn = sqlite3.connect(str(config.active_db()))
        exists = conn.execute(
            "SELECT COUNT(*) FROM merchants WHERE UPPER(merchant_name) = ?",
            (merchant.upper(),),
        ).fetchone()[0] > 0
    except Exception:
        exists = True  # fail-open if DB check itself fails
    finally:
        if conn:
            conn.close()
    if not exists:
        raise HTTPException(status_code=400,
                            detail=f"merchant_name not found in registry: {merchant}")
    engine = get_searcher().matcher.alias_engine
    learned = engine.learn(query, merchant)
    return {"learned": learned, "query": query, "merchant_name": merchant}


def _quick_match_rows(identifiers: list[str]) -> list[dict]:
    """Resolve a list of identifiers (phone / MX / TID / email / account)
    to their merchant records via the identifier-aware search.

    A row counts as MATCHED when the search hit a unique identifier with
    a high-confidence score — i.e. the identifier IS in the registry.
    """
    searcher = get_searcher()
    rows = []
    for ident in identifiers:
        res = searcher.search(ident, limit=1, min_score=0)
        best = res[0] if res else None
        rec = best.record if best else {}
        matched = bool(best and best.identifier_hit and best.overall_score >= 85)
        rows.append({
            "input": ident,
            "matched": matched,
            "matched_field": best.identifier_hit if best else "",
            "matched_value": (rec.get(best.identifier_hit, "")
                              if best and best.identifier_hit else ""),
            "best_match": rec.get("merchant_name", ""),
            "score": round(best.overall_score / 10, 1) if best else 0,
            "match_type": best.match_type if best else "Not Found",
            "email": rec.get("email", ""),
            "phone": rec.get("phone", ""),
            "tid": rec.get("tid", ""),
            "mxcode": rec.get("mxcode", ""),
            "sheet": rec.get("sheet_name", ""),
        })
    return rows


@app.post("/api/quickmatch")
def quickmatch(req: QuickMatchRequest):
    """Resolve a batch of identifiers (phones, MX codes, TIDs, emails, account
    numbers) against the registry. Unlike /api/batch (name fuzzy matching),
    this is precision-first: a record only counts as matched when the
    identifier itself was found."""
    identifiers = [i.strip() for i in req.identifiers if i.strip()][:2000]
    if not identifiers:
        raise HTTPException(status_code=400, detail="identifiers is empty")
    t0 = time.perf_counter()
    rows = _quick_match_rows(identifiers)
    elapsed = time.perf_counter() - t0

    matched = sum(1 for r in rows if r["matched"])
    emails = sum(1 for r in rows if r["email"])
    return {
        "count": len(rows),
        "elapsed_s": round(elapsed, 2),
        "matched": matched,
        "missing": len(rows) - matched,
        "emails": emails,
        "pct": round(matched / len(rows) * 100) if rows else 0,
        "rows": rows,
    }


@app.post("/api/quickmatch/export")
def quickmatch_export(req: QuickMatchRequest):
    """Resolve identifiers and export the results as an Excel workbook."""
    identifiers = [i.strip() for i in req.identifiers if i.strip()][:2000]
    if not identifiers:
        raise HTTPException(status_code=400, detail="identifiers is empty")
    _audit("export", json.dumps({"kind": "quickmatch",
                                 "count": len(identifiers)}))
    rows = _quick_match_rows(identifiers)
    df = pd.DataFrame(rows)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Quick Match", index=False)
        _style_workbook(writer.book)
    return Response(
        buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="quick_match_results.xlsx"'},
    )


class TaskRequest(BaseModel):
    text: str = ""
    intent: str = ""  # clarification choice: force this intent for the run
    remember: bool = False  # save this phrase -> intent for next time


def _log_task_request(detected: dict, result: dict) -> None:
    """Feed one executed task into the self-improvement loop (feedback.py).

    The request log powers rephrase detection (an empty-result request that
    gets re-asked) and pattern mining — logging must never break the request
    flow, so any failure is swallowed."""
    try:
        from merchant_intelligence import feedback
        feedback.log_request(
            kind="task",
            text=detected.get("raw") or "",
            intent=result.get("intent") or detected.get("intent"),
            intents=result.get("intents") or detected.get("intents") or [],
            confidence=detected.get("confidence", 0),
            identifier_count=detected.get("identifier_count", 0),
            rows=len(result.get("rows") or []),
            not_found=len(result.get("not_found") or []),
            entity_sig=feedback.entity_signature(detected),
        )
    except Exception:
        pass


@app.post("/api/task")
def task(req: TaskRequest):
    """Natural-language task interpreter (feature: paste a request).

    Detects whether the pasted text is a task (multi-line block + identifiers
    + instruction words) vs a plain merchant search. When it IS a task, plans
    and executes a step pipeline (e.g. TIDs -> MX codes -> static accounts +
    beneficiaries) and returns a render-ready table.

    Response:
      is_task: false              -> caller should run a normal /api/search
      is_task: true               -> intent, pipeline, columns, rows, not_found, summary
      needs_clarification: true   -> the request read ambiguously ("account
        details"), so NO pipeline ran. The caller renders `question` +
        `options` and re-posts with the chosen `intent` to force that run.

    req.intent: an explicit interpretation the caller already chose (from a
    clarification prompt) — detect_task forces exactly that intent.
    """
    from merchant_intelligence import tasks
    from merchant_intelligence.calibration import record as cal_record
    text = (req.text or "").strip()
    override = (req.intent or "").strip() or None
    if not text:
        raise HTTPException(status_code=400, detail="text is required")
    try:
        detected = tasks.detect_task(text, intent_override=override)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    _audit("task", json.dumps({"text": text[:300],
                               "intent": (detected or {}).get("intent")}))
    if not detected:
        return {"is_task": False, "reason": "looks like a normal search"}
    # Referential follow-up ("get the tids for the above merchant"): the
    # merchant comes from the PREVIOUS request, not this text. When the request
    # names no entity of its own, resolve the reference against the last
    # remembered context; with no context the pipeline answers with an honest
    # "no merchant found" message.
    own_entity = ((detected.get("identifier_count") or 0) > 0
                  or bool(detected.get("names"))
                  or bool(detected.get("segment")))
    if detected.get("references_previous") and not own_entity:
        tasks.inherit_reference(detected)
    summary = {k: detected.get(k) for k in
               ("intent", "intents", "identifier_count",
                "has_instruction", "multiline", "llm_refined", "confidence",
                "references_previous", "context_inherited", "key_merchants")}
    # Remember this request's entities as the follow-up context for the next
    # "the above merchant" style request. Requests without entities (a segment
    # with no merchant) never clobber the previous merchant.
    tasks.remember_entities(detected.get("identifiers"), detected.get("names"))
    # First pass (no explicit choice): ask when the request is ambiguous.
    if not override:
        clarify = tasks.suggest_clarification(text, detected)
        if clarify:
            # Remembered choice: the user saved an interpretation for this
            # exact phrase before — auto-run it, and tell the UI which one.
            auto = clarify.get("auto_pick")
            if auto:
                try:
                    forced = tasks.detect_task(text, intent_override=auto)
                except ValueError:
                    forced = None
                if forced:
                    try:
                        base = tasks.detect_task(text)
                    except ValueError:
                        base = None
                    predicted = (base["intent"] if base else
                                 detected["intent"])
                    conf = base.get("confidence", 0) if base else 0
                    # A remembered choice that matches the prediction is an
                    # ACCEPT; one that overrides it is an OVERRIDE — the
                    # calibration fitter learns which confidence bands the
                    # user corrects vs confirms. A request that is only a
                    # task BECAUSE of the pick (unforced detect_task is
                    # None) can't be a confirmation — the engine never
                    # predicted, so tag it as an override.
                    src = "accept"
                    if base is None or auto != predicted:
                        src = "override"
                    cal_record(text, predicted, conf, auto, source=src,
                               gap=tasks.top_two_gap(base or detected))
                    result = tasks.execute_task(forced)
                    _log_task_request(forced, result)
                    result["is_task"] = True
                    # Rebuild the summary from the FORCED task so the shown
                    # intent matches what actually ran (the un-forced
                    # detection would report change_details while the saved
                    # choice ran static_account).
                    result["detected"] = {k: forced.get(k) for k in
                                           ("intent", "intents",
                                            "identifier_count",
                                            "has_instruction", "multiline",
                                            "llm_refined", "confidence")}
                    result["used_preference"] = auto
                    return result
            return {
                "is_task": True,
                "needs_clarification": True,
                "question": clarify["question"],
                "options": clarify["options"],
                "detected": summary,
            }
        # Auto-routed: the engine's pick ran without a challenge — record it
        # as an accepted decision (implicit feedback for calibration). The
        # top-2 gap rides along when two+ intents scored (a race context)
        # so the gap_threshold fitter has the data it needs.
        cal_record(text, detected["intent"], detected.get("confidence", 0),
                   detected["intent"], source="auto",
                   gap=tasks.top_two_gap(detected))
    else:
        # The user picked an interpretation (clarification card / intent
        # override). Record predicted-vs-chosen so the fitter can learn
        # which confidence bands get corrected — tagged accept when the
        # user confirmed the engine's guess, override when they corrected it.
        try:
            base = tasks.detect_task(text)
        except ValueError:
            base = None
        predicted = base["intent"] if base else detected["intent"]
        conf = base.get("confidence", 0) if base else 0
        chosen = detected["intent"]
        # A request that is only a task because of the override (unforced
        # detect_task returns None) means the engine never predicted — that
        # can't be a confirmation, so it's tagged as an override.
        src = "accept"
        if base is None or chosen != predicted:
            src = "override"
        cal_record(text, predicted, conf, chosen, source=src,
                   gap=tasks.top_two_gap(base or detected))
        # "Remember my choice": persist phrase -> chosen intent so future
        # identical requests auto-run it (see suggest_clarification).
        if req.remember and override:
            try:
                from merchant_intelligence import preferences
                preferences.learn(text, override, base)
            except Exception as exc:
                logger.warning("failed to remember choice: %s", exc)
    result = tasks.execute_task(detected)
    _log_task_request(detected, result)
    result["is_task"] = True
    result["detected"] = summary
    return result


class SuggestionAction(BaseModel):
    ngram: str = ""
    intent: str = ""
    weight: int = 0


@app.get("/api/feedback/suggestions")
def feedback_suggestions():
    """Self-improvement loop status: mined pattern suggestions + outcome stats.

    Suggestions come from real corrections — clarification overrides (user
    picked a different intent than predicted) and rephrased requests (a
    request that returned nothing was re-asked with new wording). Only
    n-grams with >= 3 corroborating samples are suggested; applying writes
    them to intents.json (hot-reloaded), rejecting records them so they
    never resurface.
    """
    from merchant_intelligence import feedback
    return feedback.report()


@app.post("/api/feedback/suggestions/apply")
def suggestion_apply(req: SuggestionAction):
    """Accept a mined suggestion: write it to intents.json + hot-reload."""
    from merchant_intelligence import feedback
    spec = feedback.apply_pattern(req.ngram.strip(),
                                  (req.intent or "").strip().lower(),
                                  req.weight or None)
    if spec is None:
        raise HTTPException(status_code=400,
                            detail="unknown intent or empty phrase")
    return {"ok": True, "ngram": req.ngram.strip(),
            "intent": (req.intent or "").strip().lower(),
            "spec": spec}


@app.post("/api/feedback/suggestions/reject")
def suggestion_reject(req: SuggestionAction):
    """Reject a mined suggestion so it never resurfaces."""
    from merchant_intelligence import feedback
    feedback.reject(req.ngram.strip(), (req.intent or "").strip().lower())
    return {"ok": True}


class SynonymStatusRequest(BaseModel):
    ids: List[str] = []
    status: str = "approved"  # "approved" | "rejected"


class SynonymApplyRequest(BaseModel):
    ids: Optional[List[str]] = None  # None = apply all approved


@app.get("/api/synonyms")
def get_synonym_candidates():
    """Tier-1 WordNet proposals (design doc §4): pending/approved/rejected
    candidate phrases grouped for curation on the Rule Engine page.
    """
    from merchant_intelligence.tasks import enrichment
    return enrichment.candidates()


@app.post("/api/synonyms/propose")
def propose_synonyms():
    """Re-run the WordNet proposal stage (idempotent — statuses preserved).
    400 with an install hint when nltk/wordnet is unavailable."""
    from merchant_intelligence.tasks import enrichment
    r = enrichment.propose_candidates()
    if not r.get("ok"):
        hint = (r.get("wordnet") or {}).get("hint")
        raise HTTPException(
            status_code=400,
            detail=(r.get("reason") or "proposal failed")
                   + (f" — {hint}" if hint else ""))
    return r


@app.post("/api/synonyms/status")
def synonym_status(req: SynonymStatusRequest):
    """The curation gate: mark candidate ids approved or rejected."""
    from merchant_intelligence.tasks import enrichment
    r = enrichment.set_status(req.ids, req.status)
    if not r.get("ok"):
        raise HTTPException(status_code=400,
                            detail=r.get("reason", "bad request"))
    return r


@app.post("/api/synonyms/apply")
def apply_synonyms(req: SynonymApplyRequest):
    """Merge approved candidates into intents.json (weight-2 patterns),
    regenerating vocab.py's defaults in lockstep, appending the phrases to
    the Tier-2 exemplars and recording provenance. Hot-reloaded."""
    from merchant_intelligence.tasks import enrichment
    return enrichment.apply_approved(req.ids)


@app.get("/api/synonyms/manifest")
def synonym_manifest():
    """Provenance of every applied auto-pattern (data/auto_pattern_manifest.json)."""
    from merchant_intelligence.tasks import enrichment
    return enrichment.manifest()


class ShadowReviewLabelRequest(BaseModel):
    entry_id: str = ""
    correct: bool = True
    intent: str = ""   # optional: the intent it SHOULD have been (on a miss)
    note: str = ""


@app.get("/api/shadow/review")
def shadow_review(band: str = "all", limit: int = 100):
    """Phase-1 spot-check tool (design doc §7): shadow decisions joined with
    review labels, band-filtered, plus per-intent precision on the
    high-confidence would-act band — the band clarification labels never
    cover. This is the evidence the Phase 2 go/no-go needs."""
    from merchant_intelligence import calibration
    from merchant_intelligence.tasks import semantic
    if band not in ("all", "would_act", "would_not"):
        raise HTTPException(status_code=400,
                            detail="band must be all|would_act|would_not")
    out = semantic.review(band=band, limit=max(1, min(500, int(limit))))
    # Phase 3: per-intent fitted gates (the auto-run band's accept/override
    # evidence) so the panel can show learning progress in the same view.
    out["tier2_fit"] = calibration.fit_tier2()
    # Shadow-log health (band-independent): lets the Rule Engine chip show
    # today's entry count without opening this panel.
    out["health"] = semantic.shadow_health()
    return out


@app.post("/api/shadow/review")
def shadow_review_label(req: ShadowReviewLabelRequest):
    """Record a reviewer's verdict on one shadow entry (latest wins)."""
    from merchant_intelligence.tasks import semantic
    eid = (req.entry_id or "").strip()
    if not eid:
        raise HTTPException(status_code=400, detail="entry_id is required")
    known = {semantic.entry_id(e) for e in semantic.read_shadow()}
    if eid not in known:
        raise HTTPException(status_code=400,
                            detail="entry_id not found in the shadow log")
    return semantic.label_entry(eid, req.correct, note=req.note,
                                intent=req.intent)


@app.get("/api/audit")
def audit_endpoint(limit: int = 200, action: str = "", actor: str = ""):
    """Immutable audit trail (docs/technical-review-2026-08-original.md #1).

    Newest-first entries for every search, profile view, export, and intent
    execution, plus per-action stats. Append-only by construction — there is
    no update/delete path anywhere in merchant_intelligence/audit.py.
    """
    from merchant_intelligence import audit
    return {
        "ok": True,
        "entries": audit.recent(limit=max(1, min(1000, int(limit))),
                                action=action or None, actor=actor or None),
        "stats": audit.stats(),
        "file": str(audit._path()),
    }


@app.post("/api/task/analyze")
def task_analyze(req: TaskRequest):
    """Intent-parser debug endpoint (v2): explain WHY a request was routed
    the way it was.

    Returns every detected intent with its score / confidence / matched
    patterns, the extracted parameters (segment, names, state filter,
    presence filters, limit), and the resulting task descriptor. Use it to
    debug a mis-routed request or to render a "how the parser read this"
    panel in the UI.
    """
    from merchant_intelligence import tasks
    text = (req.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="text is required")
    _audit("task_analyze", json.dumps({"text": text[:300]}))
    try:
        return tasks.analyze(text)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.get("/api/calibration")
def get_calibration():
    """Confidence calibration status (feature: fit ask thresholds from usage).

    Returns the logged decision history summary (samples, acceptance by
    confidence band and per intent) plus the fitted thresholds that
    suggest_clarification is currently using — so the Rule Engine UI can
    show how the engine is learning to flag low-confidence requests.
    """
    from merchant_intelligence import calibration
    return {
        "stats": calibration.stats(),
        "fit": calibration.fit(),
        "params": calibration.params(),
        # Phase 3 (design doc §7): per-intent Tier-2 gates fitted from the
        # shadow-review labels — the auto-run band's accept/override evidence.
        "tier2": calibration.fit_tier2(),
    }


@app.post("/api/calibration/reset")
def reset_calibration():
    """Wipe the decision log ("start learning fresh")."""
    from merchant_intelligence import calibration
    removed = calibration.reset()
    return {"ok": True, "removed": removed,
            "stats": calibration.stats()}


@app.get("/api/preferences")
def get_preferences():
    """Saved clarification choices (phrase -> intent) for the Rule Engine UI.

    Each entry is a normalized request phrase the user answered and asked to
    remember ("account details" -> static_account). The UI lists them so a
    wrong or stale memory can be deleted.
    """
    from merchant_intelligence import preferences
    from merchant_intelligence.tasks.engine import CLARIFY_OPTIONS
    prefs = preferences.all_prefs()
    return {
        "count": len(prefs),
        "preferences": [
            {"key": k, "intent": v,
             "label": CLARIFY_OPTIONS.get(v, (v, ""))[0]}
            for k, v in sorted(prefs.items())
        ],
    }


class PreferenceForgetRequest(BaseModel):
    key: str = ""


@app.post("/api/preferences/forget")
def forget_preference(req: PreferenceForgetRequest):
    """Forget one saved interpretation (the Rule Engine delete button)."""
    from merchant_intelligence import preferences
    removed = preferences.forget((req.key or "").strip())
    return {"ok": True, "removed": removed,
            "preferences": get_preferences()}


class IntentPattern(BaseModel):
    pattern: str
    weight: int = 1


class IntentUpdateRequest(BaseModel):
    intent: str
    patterns: list[IntentPattern] = Field(default_factory=list)
    keywords: list[str] = Field(default_factory=list)
    # Per-intent typo-tolerance toggle (see vocab.INTENT_FUZZY). None keeps
    # the current value; false restricts this intent to exact regex patterns.
    fuzzy: Optional[bool] = None


@app.get("/api/intents")
def get_intents():
    """Intent config for the Rule Engine tuning UI (read-only).

    Returns the loaded intents.json (patterns + keywords + _help), where it
    is loaded from, which intents have pipelines registered, and the built-in
    defaults so the UI can offer a one-click "restore defaults".
    """
    from merchant_intelligence.tasks import vocab
    from merchant_intelligence.tasks.pipelines import _PIPELINES
    data = vocab.get_intent_config()
    return {
        "source": vocab.intents_source(),
        "help": data.get("_help", ""),
        "intents": data.get("intents", {}),
        "defaults": vocab.default_intent_specs(),
        "pipelines": sorted(_PIPELINES),
        "name_capable": sorted(vocab.NAME_CAPABLE_INTENTS),
        "chainable": vocab.CHAINABLE,
    }


@app.put("/api/intents")
def update_intent(req: IntentUpdateRequest):
    """Update one intent's patterns/keywords, persist to intents.json, and
    hot-reload the engine so the change applies immediately (no restart)."""
    from merchant_intelligence.tasks import vocab
    intent = (req.intent or "").strip().lower()
    data = vocab.get_intent_config()
    if intent not in (data.get("intents") or {}):
        raise HTTPException(status_code=404, detail=f"unknown intent: {intent}")
    spec = {
        "patterns": [{"pattern": p.pattern, "weight": p.weight} for p in req.patterns],
        "keywords": [k.strip() for k in req.keywords if k.strip()],
    }
    # Preserve the fuzzy toggle through a save: the UI always sends it (the
    # editor's toggle), but a partial client must not silently flip it off.
    if req.fuzzy is not None:
        spec["fuzzy"] = req.fuzzy
    else:
        cur = (data.get("intents") or {}).get(intent)
        if isinstance(cur, dict) and isinstance(cur.get("fuzzy"), bool):
            spec["fuzzy"] = cur["fuzzy"]
    errors = vocab.validate_intent_spec(spec)
    if errors:
        raise HTTPException(status_code=400, detail="; ".join(errors))
    try:
        vocab.save_intent_config(intent, spec)
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"could not write config: {exc}")
    return {
        "ok": True,
        "intent": intent,
        "hot_reloaded": True,
        "intents": vocab.get_intent_config().get("intents", {}).get(intent),
    }


class SettingsUpdateRequest(BaseModel):
    decisive_match_threshold: Optional[float] = Field(
        None, ge=0.0, le=100.0)
    # Tier-2 semantic rollout state: "off" | "shadow" | "enabled"
    # (validated in the handler — same choices as settings._MODE_SPEC).
    semantic_tier_mode: Optional[str] = None


@app.get("/api/settings")
def get_settings():
    """Runtime engine settings for the Rule Engine tuning UI (read-only).

    Returns every tunable knob's resolved value, its built-in default, the
    precedence source (env var > data/engine_settings.json > default) and
    the valid range — so the UI can render a proper tuning control.
    """
    from merchant_intelligence import settings as engine_settings
    return {
        "ok": True,
        "settings": engine_settings.all_settings(),
        "file": str(engine_settings._path()),
    }


@app.put("/api/settings")
def update_settings(req: SettingsUpdateRequest):
    """Update runtime engine settings, persist to data/engine_settings.json
    and apply immediately (no restart).

    Settings are read on every use (profile.py reads the threshold per
    build), so a save here is hot-reloaded the moment the next request runs.
    Only the knobs supplied are changed; missing fields keep their value.
    """
    from merchant_intelligence import settings as engine_settings
    current = engine_settings.load()
    if req.decisive_match_threshold is not None:
        current["decisive_match_threshold"] = req.decisive_match_threshold
    if req.semantic_tier_mode is not None:
        if req.semantic_tier_mode not in ("off", "shadow", "enabled"):
            raise HTTPException(
                status_code=400,
                detail="semantic_tier_mode must be one of: off, shadow, enabled")
        current["semantic_tier_mode"] = req.semantic_tier_mode
    engine_settings.save(current)
    return {
        "ok": True,
        "hot_reloaded": True,
        "settings": engine_settings.all_settings(),
        "file": str(engine_settings._path()),
    }


@app.delete("/api/settings")
def reset_settings():
    """Delete data/engine_settings.json so every knob falls back to its
    built-in default (the 'Reset to defaults' button on the Rule Engine)."""
    from merchant_intelligence import settings as engine_settings
    try:
        path = engine_settings._path()
        if path.exists():
            path.unlink()
    except OSError:
        pass
    return {
        "ok": True,
        "settings": engine_settings.all_settings(),
        "file": str(engine_settings._path()),
    }


# Header -> row key mapping for task export (mirrors the frontend's
# TASK_COLUMN_KEYS) so every pipeline — static_account, email/phone/mx,
# profile, change_details, segment, count, duplicates, summary — exports
# its full column set instead of the old hard-coded subset.
_TASK_EXPORT_KEY_BY_HEADER = {
    "TID": "tid", "Merchant": "merchant", "Merchant Name": "merchant",
    "MX Code": "mxcode", "Static Account Number": "static_acc_no",
    "Beneficiary": "beneficiary", "Payable Code": "payable_code",
    "Payable": "payable", "Alias": "alias", "Bank": "bank",
    "Status": "status", "Identifier": "identifier", "Phone": "phone",
    "Email": "email", "Slip Header": "slip_header", "Source": "sheet",
    "Sheet": "sheet", "Account Name": "account_name",
    "Account Number": "account_number", "Contact": "contact",
    "Address": "address", "State": "state", "Onboarded": "onboarded",
    "Row": "row", "Metric": "metric", "Count": "count", "Value": "value",
    "Rows": "rows", "Sources": "sources", "Current Account": "current_acc",
    "Current Bank": "current_bank", "Changed": "change_detected",
    # change-details old/new pairs are stored under the exact header text
    "Old Bank Acc No": "Old Bank Acc No", "New Bank Acc No": "New Bank Acc No",
    "Old Bank Code": "Old Bank Code", "New Bank Code": "New Bank Code",
    "Old Address": "Old Address", "New Address": "New Address",
    "Old Account Name": "Old Account Name", "New Account Name": "New Account Name",
}

# Columns whose row key varies by pipeline. Resolution order mirrors the
# frontend: prefer the row's own key, fall back through the alternates.
#   State  — segment rows carry 'state'; profile rows store state under 'bank'
#   Source — segment rows use 'source'; every other pipeline uses 'sheet'
_TASK_EXPORT_ALTERNATES = {
    "State": ["state", "bank"],
    "Source": ["source", "sheet"],
}


def _key_merchants_for(name: str) -> list:
    """Key-merchant roots a merchant name belongs to ([] if none).

    Same engine the Search page badge uses (SearchResult._key_merchants),
    so the Similar/Related panel and Batch rows show exactly what the
    search badge shows for the same name. Lazy import — api.py must never
    import the tasks package at module load.
    """
    try:
        from merchant_intelligence.tasks.parser import key_merchant_matches
        return list(key_merchant_matches(name or ""))
    except Exception as exc:  # badge is a nicety — never break the API row
        logger.warning("key_merchant_matches failed for %r: %s", name, exc)
        return []


def _task_export_snake(header: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", header.lower())


# Shared workbook styling for every export endpoint: dark-blue bold header,
# thin borders, zebra striping, frozen header row, autofilter, auto column
# widths (capped so wide free-text like addresses don't explode the sheet)
# and wrap for long values. Matches the polished look of the standalone
# export script so every downloaded xlsx feels the same.
_HEADER_FILL = "1F4E78"
_ZEBRA_FILL = "EAF1F8"
_MAX_COL_WIDTH = 55
_MIN_COL_WIDTH = 9


def _style_workbook(wb) -> None:
    """Apply the standard export style to every worksheet in a workbook.

    Call AFTER writing the DataFrames (inside the pd.ExcelWriter context)
    with `writer.book` so the rows exist before styling runs.
    """
    from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    zebra = PatternFill("solid", fgColor=_ZEBRA_FILL)
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        max_col = ws.max_column
        max_row = ws.max_row
        if max_row < 1 or max_col < 1:
            continue
        # Auto column widths from content (capped; header text counts too).
        for c in range(1, max_col + 1):
            longest = 0
            for r in range(1, max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                longest = max(longest, len(str(v)))
            ws.column_dimensions[get_column_letter(c)].width = max(
                _MIN_COL_WIDTH, min(_MAX_COL_WIDTH, longest + 3))
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if r == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center",
                                               vertical="center",
                                               wrap_text=True)
                else:
                    if r % 2 == 0:
                        cell.fill = zebra
                    v = cell.value
                    if v is not None and len(str(v)) > 40:
                        cell.alignment = Alignment(vertical="top",
                                                   wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"


def _task_export_frame(result):
    """Result rows + columns -> export-ready list of dicts (human headers).

    Resolves each column's row key the same way the frontend does (mapped
    key, exact label for change-pair rows, snake_case, per-column alternates),
    then appends any extra row fields the pipeline produced so nothing is lost.
    """
    columns = result.get("columns") or []
    rows = result.get("rows") or []
    if not rows:
        return []
    used = set()
    for c in columns:
        key = _TASK_EXPORT_KEY_BY_HEADER.get(c) or _task_export_snake(c)
        used.add(key)
        used.add(c)
        used.update(_TASK_EXPORT_ALTERNATES.get(c, []))
    extra_keys = list(dict.fromkeys(k for r in rows for k in r if k not in used))
    out = []
    for r in rows:
        row = {}
        for c in columns:
            key = _TASK_EXPORT_KEY_BY_HEADER.get(c) or _task_export_snake(c)
            val = r.get(key)
            if val in (None, ""):
                val = r.get(c)  # exact-label keys (change old/new pairs)
            if val in (None, ""):
                val = r.get(_task_export_snake(c))
            if val in (None, ""):
                for alt in _TASK_EXPORT_ALTERNATES.get(c, []):
                    v = r.get(alt)
                    if v not in (None, ""):
                        val = v
                        break
            row[c] = "" if val is None else val
        for k in extra_keys:
            if k not in row:
                row[k] = "" if r.get(k) is None else r.get(k)
        out.append(row)
    return out


@app.post("/api/task/export")
def task_export(req: TaskRequest):
    """Interpret a pasted request, execute it, and export the result table as
    an Excel workbook (used by the Export button on task results)."""
    from merchant_intelligence import tasks
    text = (req.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="text is required")
    try:
        detected = tasks.detect_task(text)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if not detected:
        raise HTTPException(status_code=400, detail="not a task - nothing to export")
    _audit("export", json.dumps({"kind": "task", "text": text[:300],
                                 "intent": detected.get("intent")}))
    result = tasks.execute_task(detected)
    rows = result.get("rows", [])
    columns = result.get("columns", [])
    if not columns or not rows:
        raise HTTPException(status_code=400, detail="task produced no rows to export")
    df = pd.DataFrame(_task_export_frame(result))
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Task Results", index=False)
        pd.DataFrame({"Note": [result.get("summary", "")]}).to_excel(
            writer, sheet_name="Summary", index=False)
        _style_workbook(writer.book)
    return Response(
        buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="task_results.xlsx"'},
    )


@app.post("/api/batch")
def batch(req: BatchRequest):
    merchants = [m.strip() for m in req.merchants if m.strip()][:1000]
    if not merchants:
        raise HTTPException(status_code=400, detail="merchants is empty")
    _audit("batch", json.dumps({"count": len(merchants)}))
    t0 = time.perf_counter()
    rows = _run_batch(merchants)
    elapsed = time.perf_counter() - t0

    found = sum(1 for r in rows if r["score"] >= 5.0)
    emails = sum(1 for r in rows if r["email"])
    return {
        "count": len(rows),
        "elapsed_s": round(elapsed, 2),
        "found": found,
        "missing": len(rows) - found,
        "emails": emails,
        "pct": round(found / len(rows) * 100) if rows else 0,
        "rows": rows,
    }


@app.post("/api/batch/export")
def batch_export(req: BatchRequest):
    merchants = [m.strip() for m in req.merchants if m.strip()][:1000]
    rows = _run_batch(merchants)
    _audit("export", json.dumps({"kind": "batch", "count": len(rows)}))
    df = pd.DataFrame(rows)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Batch Search", index=False)
        _style_workbook(writer.book)
    return Response(
        buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="batch_search_results.xlsx"'},
    )


@app.get("/api/quality")
def quality():
    """Data quality scan of the registry (reuses data_quality.run_quality)."""
    from data_quality import run_quality
    q = run_quality()
    total = q["total"] or 0
    return {
        "total": total,
        "sheets": q.get("sheets", {}),
        "missing": q.get("missing", {}),
        "code_names": q.get("code_names", 0),
        "orphans": q.get("orphans", 0),
        "duplicate_tids": q["duplicate_tids"].to_dict(orient="records"),
        "mx_multiname": q["mx_multiname"].to_dict(orient="records"),
    }


@app.post("/api/quality/export")
def quality_export():
    """Export the data quality report as an Excel workbook."""
    from data_quality import run_quality
    _audit("export", json.dumps({"kind": "quality"}))
    q = run_quality()
    total = q["total"] or 0
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame({
            "Metric": ["Total records", "Code names", "Orphans"]
            + [f"Missing {f}" for f in q["missing"]],
            "Value": [q["total"], q["code_names"], q["orphans"]]
            + [q["missing"][f] for f in q["missing"]],
        }).to_excel(writer, sheet_name="Summary", index=False)
        q["duplicate_tids"].to_excel(writer, sheet_name="Duplicate TIDs", index=False)
        q["mx_multiname"].to_excel(writer, sheet_name="MX Multi-Name", index=False)
        pd.DataFrame(
            [{"Sheet": s, "Records": n} for s, n in q.get("sheets", {}).items()]
        ).to_excel(writer, sheet_name="Sheets", index=False)
        _style_workbook(writer.book)
    return Response(
        buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="data_quality_report.xlsx"'},
    )


@app.post("/api/reconcile")
def reconcile_endpoint(req: BatchRequest):
    """Reconcile a merchant list into verified matches + recovered assets."""
    from reconcile import reconcile as run_reconcile
    merchants = [m.strip() for m in req.merchants if m.strip()][:1000]
    if not merchants:
        raise HTTPException(status_code=400, detail="merchants is empty")
    _audit("reconcile", json.dumps({"count": len(merchants)}))
    t0 = time.perf_counter()
    report = run_reconcile(merchants, top_n=3)
    elapsed = time.perf_counter() - t0

    def rows(name):
        df = report[name]
        return df.to_dict(orient="records") if df is not None and not df.empty else []

    matches = rows("matches")
    not_found = rows("not_found")
    emails = rows("emails")
    contacts = rows("contacts")
    return {
        "count": len(merchants),
        "elapsed_s": round(elapsed, 2),
        "found": len(matches),
        "missing": len(not_found),
        "emails": len(emails),
        "contacts": len(contacts),
        "pct": round(len(matches) / len(merchants) * 100) if merchants else 0,
        "matches": matches,
        "not_found": not_found,
        "emails_rows": emails,
        "contacts_rows": contacts,
        "summary": rows("summary"),
        "all_results": rows("all_results"),
    }


@app.post("/api/brief")
def brief(req: ProfileRequest):
    """LLM investigation brief for a merchant fragment (feature #6).

    Builds the 360° profile for the query (any fragment — name, email, phone,
    MX code, TID, account number…) then produces a natural-language
    investigation dossier. Uses an LLM when LLM_API_KEY is configured
    (any OpenAI-compatible endpoint via LLM_BASE_URL/LLM_MODEL); otherwise
    falls back to a deterministic offline template brief.
    """
    from merchant_intelligence.brief import build_brief, llm_available
    query = req.query.strip()
    if not query:
        raise HTTPException(status_code=400, detail="query is required")
    _audit("brief", json.dumps({"query": query}))
    profile = get_profiler().build(query, max_members=req.max_members)
    out = build_brief(profile)
    out["query"] = query
    out["llm_configured"] = llm_available()
    return out


@app.get("/api/selfimprove")
def selfimprove_status():
    """Last alias-free harness run + current baseline (feature #10).

    Returns the stored self-improve report (data/self_improve_report.json) and
    baseline (data/alias_free_baseline.json) so the frontend can show engine
    health. Empty dicts when the harness has not run yet.
    """
    import json
    out = {"report": None, "baseline": None}
    for key, name in (("report", "self_improve_report.json"),
                      ("baseline", "alias_free_baseline.json")):
        try:
            path = config.DATA_DIR / name
            if path.exists():
                out[key] = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            out[key] = None
    return out


@app.post("/api/reconcile/export")
def reconcile_export(req: BatchRequest):
    """Run reconciliation and export the full report as Excel."""
    from reconcile import reconcile as run_reconcile
    merchants = [m.strip() for m in req.merchants if m.strip()][:1000]
    if not merchants:
        raise HTTPException(status_code=400, detail="merchants is empty")
    report = run_reconcile(merchants, top_n=3)
    buffer = BytesIO()
    sheet_names = {
        "summary": "Summary", "matches": "Matches", "not_found": "Not Found",
        "all_results": "All Results", "emails": "Emails", "contacts": "Contacts",
    }
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for key, name in sheet_names.items():
            df = report[key]
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=name, index=False)
            else:
                pd.DataFrame({"Note": ["No records"]}).to_excel(writer, sheet_name=name, index=False)
        _style_workbook(writer.book)
    return Response(
        buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Merchant_Reconciliation_Report.xlsx"'},
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)
