"""api_shared.py — shared infrastructure for the api_routes package.

Extracted from api.py during the roadmap #3 router split so handlers live in
domain routers (api_routes/) while everything they share — the audit helper,
lazy singletons, identifier-aware search, and the pydantic request models —
lives here. api.py imports this module, mounts the routers, and re-exports
every handler/model so `import api; api.search(...)` keeps working.
"""

import json
import logging
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
# CLI tools (report.py, data_quality.py, reconcile.py) live in scripts/
_SCRIPTS_DIR = PROJECT_ROOT / "scripts"
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from fastapi import HTTPException  # noqa: E402  (after sys.path setup)
from merchant_intelligence import config  # noqa: E402
from merchant_intelligence.entity import EntityResolver  # noqa: E402
from merchant_intelligence.fuzzy import (  # noqa: E402
    damerau_levenshtein_similarity,
    token_sort_ratio,
)


def _audit(action: str, scope: Optional[str] = None,
           detail: Optional[str] = None) -> None:
    """Best-effort audit-trail append (roadmap #1) — never breaks the
    request; audit.record() already swallows its own failures. Actor comes
    from the auth middleware context (the session username) or "local"."""
    try:
        from merchant_intelligence import audit
        from merchant_intelligence import auth
        audit.record(action, scope=scope, detail=detail,
                     actor=auth.current_actor())
    except Exception:  # noqa: BLE001
        pass


_searcher = None
_resolver = None
_profiler = None


def _multi_identifier_query(query: str) -> Optional[List[str]]:
    """Return the identifier values when a query pastes 2+ of them at once.

    ``"2ISW2587 2ISW2586"`` (two TIDs) or ``"MX183544 MX183545"`` is a
    batch-resolve request: each identifier must be searched on its own and
    the mentions merged — treating the whole string as one search blob makes
    FTS fuzzy-match unrelated rows (e.g. UBTH records matching on loose
    token overlap). A single identifier or a plain name query returns None
    and searches normally.
    """
    try:
        from merchant_intelligence.tasks.parser import parse_identifiers
        ident = parse_identifiers(query)
        values = [v for vals in ident.values() for v in vals]
        values = list(dict.fromkeys(values))  # dedupe, preserve order
        if len(values) >= 2:
            return values
    except Exception:
        pass
    return None


def _search_with_multi(query: str, fetch: int, min_score: float) -> list:
    """Search, OR-ing the identifiers when the query pastes 2+ of them.

    Every returned SearchResult keeps its own match metadata (matched_field /
    matched_value) from its identifier's search, and rows are deduped by
    record id so a value stored in two columns never double-counts.
    """
    searcher = get_searcher()
    multi = _multi_identifier_query(query)
    if not multi:
        return searcher.search(query, limit=fetch, min_score=min_score)
    merged: Dict[Any, Any] = {}
    for v in multi:
        # Each identifier is searched on its own; only exact/high-confidence
        # hits count so the loose fuzzy tier (UBTH/terminal-serial noise that
        # matches a bare TID's digits) never pollutes a batch-resolve result.
        for r in searcher.search(v, limit=min(fetch, 25), min_score=min_score):
            if r.match_type not in ("Exact Match", "High Confidence"):
                continue
            rid = r.record.get("id")
            if rid is not None and rid not in merged:
                merged[rid] = r
    return list(merged.values())


def get_searcher():
    from merchant_intelligence import MerchantSearch
    global _searcher
    if _searcher is None:
        _searcher = MerchantSearch()
    return _searcher


def get_resolver() -> EntityResolver:
    global _resolver
    if _resolver is None:
        _resolver = EntityResolver()
    return _resolver


def get_profiler():
    """Return the cached MerchantProfile singleton (avoids re-opening the
    DatabaseManager/EntityResolver/Matcher connections per request)."""
    global _profiler
    if _profiler is None:
        from merchant_intelligence.profile import MerchantProfile
        _profiler = MerchantProfile()
    return _profiler


def _key_merchants_for(name: str) -> list:
    """Key-merchant roots a merchant name belongs to ([] if none).

    Same engine the Search page badge uses (SearchResult._key_merchants),
    so the Similar/Related panel and Batch rows show exactly what the
    search badge shows for the same name. Lazy import — api modules must
    never import the tasks package at module load.
    """
    try:
        from merchant_intelligence.tasks.parser import key_merchant_matches
        return list(key_merchant_matches(name or ""))
    except Exception as exc:  # badge is a nicety — never break the API row
        logger.warning("key_merchant_matches failed for %r: %s", name, exc)
        return []


def _quick_match_rows(identifiers: list[str]) -> list[dict]:
    """Resolve a list of identifiers (phone / MX / TID / email / account)
    to their merchant records via the identifier-aware search.

    A row counts as MATCHED when the search hit a unique identifier with
    a high-confidence score — i.e. the identifier IS in the registry.
    """
    searcher = get_searcher()
    rows = []
    for ident in identifiers:
        res = searcher.search(ident, limit=1, min_score=0)
        best = res[0] if res else None
        rec = best.record if best else {}
        matched = bool(best and best.identifier_hit and best.overall_score >= 85)
        rows.append({
            "input": ident,
            "matched": matched,
            "matched_field": best.identifier_hit if best else "",
            "matched_value": (rec.get(best.identifier_hit, "")
                              if best and best.identifier_hit else ""),
            "best_match": rec.get("merchant_name", ""),
            "score": round(best.overall_score / 10, 1) if best else 0,
            "match_type": best.match_type if best else "Not Found",
            "email": rec.get("email", ""),
            "phone": rec.get("phone", ""),
            "tid": rec.get("tid", ""),
            "mxcode": rec.get("mxcode", ""),
            "sheet": rec.get("sheet_name", ""),
        })
    return rows


def _log_task_request(detected: dict, result: dict) -> None:
    """Feed one executed task into the self-improvement loop (feedback.py).

    The request log powers rephrase detection (an empty-result request that
    gets re-asked) and pattern mining — logging must never break the request
    flow, so any failure is swallowed."""
    try:
        from merchant_intelligence import feedback
        feedback.log_request(
            kind="task",
            text=detected.get("raw") or "",
            intent=result.get("intent") or detected.get("intent"),
            intents=result.get("intents") or detected.get("intents") or [],
            confidence=detected.get("confidence", 0),
            identifier_count=detected.get("identifier_count", 0),
            rows=len(result.get("rows") or []),
            not_found=len(result.get("not_found") or []),
            entity_sig=feedback.entity_signature(detected),
        )
    except Exception:
        pass


# ── shared workbook styling (every export endpoint uses this) ────────────
# Dark-blue bold header, thin borders, zebra striping, frozen header row,
# autofilter, auto column widths (capped so wide free-text like addresses
# don't explode the sheet) and wrap for long values. Matches the polished
# look of the standalone export script so every downloaded xlsx feels the
# same.
_HEADER_FILL = "1F4E78"
_ZEBRA_FILL = "EAF1F8"
_MAX_COL_WIDTH = 55
_MIN_COL_WIDTH = 9


def _style_workbook(wb) -> None:
    """Apply the standard export style to every worksheet in a workbook.

    Call AFTER writing the DataFrames (inside the pd.ExcelWriter context)
    with `writer.book` so the rows exist before styling runs.
    """
    from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    zebra = PatternFill("solid", fgColor=_ZEBRA_FILL)
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        max_col = ws.max_column
        max_row = ws.max_row
        if max_row < 1 or max_col < 1:
            continue
        # Auto column widths from content (capped; header text counts too).
        for c in range(1, max_col + 1):
            longest = 0
            for r in range(1, max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                longest = max(longest, len(str(v)))
            ws.column_dimensions[get_column_letter(c)].width = max(
                _MIN_COL_WIDTH, min(_MAX_COL_WIDTH, longest + 3))
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if r == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center",
                                               vertical="center",
                                               wrap_text=True)
                else:
                    if r % 2 == 0:
                        cell.fill = zebra
                    v = cell.value
                    if v is not None and len(str(v)) > 40:
                        cell.alignment = Alignment(vertical="top",
                                                   wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"


# ── pydantic request models (shared by all routers) ──────────────────────


class SearchRequest(BaseModel):
    query: str
    limit: int = Field(20, ge=1, le=100)
    min_score: float = Field(0.0, ge=0.0, le=100.0)
    offset: int = Field(0, ge=0)


class BatchRequest(BaseModel):
    merchants: list[str] = Field(default_factory=list)


class LearnRequest(BaseModel):
    query: str
    merchant_name: str


class EntityRequest(BaseModel):
    query: str
    depth: int = Field(2, ge=1, le=3)
    max_nodes: int = Field(120, ge=5, le=300)


class QuickMatchRequest(BaseModel):
    identifiers: list[str] = Field(default_factory=list)


class ProfileRequest(BaseModel):
    query: str
    max_members: int = Field(200, ge=1, le=1000)


class CompareRequest(BaseModel):
    query_a: str
    query_b: str
    max_members: int = Field(200, ge=1, le=1000)


class TaskRequest(BaseModel):
    text: str = ""
    intent: str = ""  # clarification choice: force this intent for the run
    remember: bool = False  # save this phrase -> intent for next time


class CopilotRequest(BaseModel):
    """Merchant Copilot (roadmap #4): a compound investigation request that
    gets decomposed into an ordered, re-runnable plan of deterministic
    steps. use_llm lets the caller force the rule-engine decomposition
    (deterministic) or allow the LLM to propose the plan when configured."""
    text: str = ""
    use_llm: bool = True
