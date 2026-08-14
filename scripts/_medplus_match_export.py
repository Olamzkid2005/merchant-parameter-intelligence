"""Match pasted Medplus store addresses against data/Medplus.xlsx ONLY and export a styled Excel file.

Output : data/medplus_tids.xlsx  (bold header, autofilter, frozen top row, status colouring)
"""
import sys, re
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.path.insert(0, 'scripts'); sys.path.insert(0, '.')

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from merchant_intelligence.fuzzy import token_sort_ratio

ADDRESSES = [
    "MEDPLUS JABI LAKE, 256A BOKA SOKORO WAY, JABI DISTRICT, CADASTRAL ZONE 304, F.C.T, ABUJA.",
    "MEDPLUS SHOP B, GROUND FLOOR, MOBOLAJI JOHNSON RAILWAY STATION, LAGOS.",
    "MEDPLUS STUDIO 24, PLOT 100, 3RD AVENUE GWARINPA, ABUJA",
    "MEDPLUS BLENCO SUPERMARKET LEKKI EPE EXPRESSWAY BY EPUTU BUS STOP IBEJU LEKKI, LAGOS STATE",
    "MEDPLUS 31A ISAAC JOHN STREET VICTORIA ISLAND LAGOS",
    "MEDPLUS ADENIRAN OGUNSANYA SHOPING MALL, SURULERE LAGOS",
    "MEDPLUS SHOP 33A, ASABA MALL, OKPANAM, ROAD, ASABA, DELTA.",
    "MEDPLUS SHOP 65/66 IBATORO ROAD, AKURE.",
    "MEDPLUS 78, NANA PLAZA AMINU-KANO CRESCENT WUSE II ABUJA.",
    "ENUGU RETAIL CENTRE LTD, OFF PRESIDENTIAL ROAD OPP MICHEAL OKPARA ENUGU",
    "MEDPLUS PLOT 2 ADMIRALTY WAY LEKKI PHASE 1 LAGOS",
    "CHEVRON DRIVE CHEVRON",
    "MEDPLUS THE PALMS, 1 BIS WAY, LEKKI, LAGOS.",
    "MEDPLUS SHOP B4 & B5, CAPITAL GARDEN MALL, ORCHID ROAD, ETI-OSA LOCAL GOVT LAGOS",
    "MEDPLUS 15 ODUDUWA WAY, IKEJA LAGOS",
    "MEDPLUS ABS AWKA ROAD BY 7 PARK ROAD ONITSHA ANAMBRA STATE",
    "MEDPLUS NOVARE MALL, CNR OF LEKKI EXP/WAY AND MONASTERY RD, SANGOTEDO, LEKKI, LAGOS.",
    "CABANA AREA, PRESIDENTIAL HOTEL, NO 1 BIRABI STREET, PORT HARCOURT - ABA EXPRESS, PORT HARCOURT, RIVERS",
    "MEDPLUS ENUGU RETAIL CENTER LIMITED, NKPOKITI ROAD, OFF PRESIDENTIAL ROAD, OPP OKPARA SQUARE, ENUGU.",
    "MEDPLUS 22 ROAD TASTY FRIED CHICKEN BUILDING FESTAC",
    "MEDPLUS WORLD BANK OWERRI IMO STATE",
    "MEDPLUS SHOP 23 CIRCLE MALL JAKANDE ROUNDABOUT LEKKI EPE EXPRESSWAY LEKKI, LAGOS STATE",
    "MEDPLUS 38 SALAMI SUAIBU STREET PADRO ROAD, PALMGROOVE, LAGOS NIGERIA",
    "SIMBIAT IKEJA MALL, 22, SIMBIAT ABIOLA WAY, IKEJA LAGOS",
    "MEDPLUS 47, CALABAR ROAD, CALABAR.",
    "MEDPLUS BERA ESTATE 20B UBA ROAD, CHEVRON DRIVE, LEKKI LAGOS",
    "MEDPLUS 107, ALLEN AVENUE, IKEJA LAGOS",
    "MEDPLUS 193, BRITISH AMERICAN JUNCTION, JOS, PLATEAU STATE",
    "58, IGBOGBO IKORODU ROAD, OPPOSITE ORIWU MODEL COLLEGE, IKORODU LAGOS",
    "MEDPLUS GREENVILLE PLAZA, 19 CMD ROAD IKOSI-KETU, LAGOS.",
    "57 ADEKUNLE BANJO STREET OFF CMD ROAD MAGODO",
    "261, HERBERT MACAULEY WAY YABA",
    "MEDPLUS PRICELESS MALL, PLOT C9 OKIGWE ROAD BY ORJI FLYOVER, OWERRI IMO STATE",
    "MEDPLUS CHEVRON DRIVE CHEVRON",
    "MEDPLUS MALL 55, KUMASI CRESCENT, WUSE 2, ABUJA",
    "1 AZIKIWE ROAD PORTHARCOURT",
    "MEDPLUS OASIS CENTER, MOB0LAJI BANK ANTHONY WAY OPPOSITE THE POLICE COLLEGE, IKEJA",
    "MEDPLUS QMB MART BLOCK, 138 PLOT 8 LEKKI SCHEME 1, LAGOS",
    "MEDPLUS 45 SAKA TINUBU STREET VICTORIA ISLAND LAGOS",
    "MEDPLUS GREENVILLE MALL, NO 10, ADEOLA ODEKU STREET, VICTORIA ISLAND",
    "10, LALUBU STREET, OKE ILEWO ABEOKUTA, OGUN",
    "MEDPLUS NEW INTERNATIONAL TERMINAL OF MMIA, IKEJA, LAGOS.",
    "39 LEKKI ESTATE ROAD WITHIN BELA VISTA ESTATE FREEDOM WAY LEKKI LAGOS",
    "MEDPLUS LACIUDAD MALL BLOCK XXVI PLOT 1 LEKKI EPE EXPRESSWAY IKOTA LAGOS",
    "MEDPLUS OLD GARKI ROAD, UMUAHIA, ABIA.",
    "FRESHFORTE, Plot C5, Block 12e, Admiralty Way, Lekki Phase 1, Lagos",
    "MEDPLUS 5 ORIWU STREET BESIDE PETROCAM FILLING STATION LEKKI PHASE 1 LAGOS",
    "PLOT 1 & 3 GIDAN KAJI ZUNGERU BY AIRPORT ROAD, KANO STATE",
    "MEDPLUS 2A ADEBAYO DOHERTY STREET, LEKKI PHASE 1",
    "MEDPLUS 113, OGUNLANA DRIVE SURULERE LAGOS",
    "6D PRICE ALABA ONIRU, VICTORIA ISLAND, LAGOS STATE (FOURPOINT BY SHERATON HOTEL ONIRU)",
    "MEDPLUS 45A, OGUDU ROAD LAGOS",
    "31, KEFFI STREET, IKOYI, LAGOS",
    "MEDPLUS NOVARE GATEWAY MALL, AIRPORT ROAD, MUSA YAR'ADUA EXPRESSWAY LUGBE",
    "MEDPLUS 53 PUMP AND FELL ALONG ADO ROAD, AJAH",
    "MEDPLUS 52 ABENI PLAZA LIGALI AYORINDE, BESIDE KFC",
    "MEDPLUS PLOT 2 ADMIRALTY WAY LEKKI PHASE 1 LAGOS",
    "MEDPLUS 22 WAFF ROAD, OFF AHMADU BELLO WAY, CITY CENTRE, KADUNA.",
    "MEDPLUS MARINA LAGOS ISLAND, LAGOS STATE",
    "ADO BAYERO MALL MALL, ZOO ROAD, KANO",
    "MEDPLUS 168 AWOLOWO ROAD, IKOYI LAGOS",
    "MEDPLUS SHOP 38 BENIN CITY MALL SAPELE ROAD EDO STATE",
    "MEDPLUS OSUN MALL, OLAIYA GBONGAN ROAD, OSOGBO, OSUN.",
    "MEDPLUS KM 45 LEKKI EPE EXPRESSWAY SANGOTEDO, LAGOS",
    "THE PALMS, OTA FORMER GATEWAY HOTEL PREMISES SANGO OTTA",
    "MEDPLUS PHASE 2, GENESIS CENTER, 39 TOMBIA ST, G.R.A PORT HARCOURT.",
    "NO 11 BAYO KUKU ROAD IKOYI LAGOS",
    "MEDPLUS PLOT 232, CADASTRAL ZONE B4, DUTSE, ABUJA.",
    "MEDPLUS 52, OPEBI ROAD, SALVATION BUS STOP, IKEJA, LAGOS",
    "MEDPLUS MCC OWERRI IMO STATE",
    "MEDPLUS 22 ROAD, GROOVE MALL FESTAC, LAGOS",
    "MEDPLUS 26, ADEOLA HOPEWELL STREET, VI LAGOS",
    "MEDPLUS KWARA MALL, ILORIN KWARA STATE ILORIN",
    "MEDPLUS AIRFORCE BASE JUNCTION, AIRPORT BASE ILORIN",
    "MEDPLUS 1004 ADMIN BLOCK HOUSING ESTATE, V.I, LAGOS.",
    "MEDPLUS 37, GLOVER ROAD, IKOYI, LAGOS.",
    "MEDPLUS ATLANTIC MALL, PLOT 482, SHOP (8 & 9) CADASTRAL ZONE B05, OBAFEMI AWOLOWO EXPRESSWAY, UTAKO ABUJA",
    "MEDPLUS 9 KEFFI STREET OFF AWOLOWO ROAD IKOYI, LAGOS STATE",
    "MEDPLUS ADO BAYERO MALL MALL, ZOO ROAD, KANO",
    "DELTA MALL EFFURUN",
    "MEDPLUS 2 ADMIRALTY ROAD, LEKKI PHASE 1, LAGOS",
    "OPEYEMI BAMIDELE FREEDOM WAY LEKKI LAGOS",
    "MEDPLUS THE BLOC, 70 KUSENLA ROAD, IKATE ELEGUSHI, OFF LEKKI-EPE EXPRESSWAY LEKKI LAGOS",
    "MEDPLUS REV OGUNBIYI STREET OFF OBA AKINJOBI WAY GRA IKEJA",
    "CABANA AREA, PRESIDENTIAL HOTEL, NO 1 BIRABI STREET, PORT HARCOURT - ABA EXPRESS, PORT HARCOURT, RIVERS",
    "MEDPLUS NO 12 FOLA OSIBO LEKKI PHASE 1 LAGOS",
    "MEDPLUS SHOP 31A, DELTA MALL, SHOPRITE, WARRI, DELTA.",
    "MEDPLUS SHOP 31A, DELTA MALL, SHOPRITE, WARRI, DELTA.",
    "THE PALMS, RINGROAD IBADAN OYO STATE IBADAN NIGERIA",
    "MEDPLUS 57 ADEKUNLE BANJO STREET OFF CMD ROAD MAGODO",
    "MEDPLUS NOVARE CENTRAL OFFICE PARK, PLOT 502 DALABA STREET, WUSE ZONE 5, ABUJA",
    "MEDPLUS 69, ADMIRALTY WAY, LEKKI PHASE 1, LAGOS",
    "MEDPLUS OGUDU CITY MALL, 175 OGUDU ROAD, OGUDU, LAGOS",
    "5, BOMPAI ROAD, FAGGAE KANO",
    "MEDPLUS NO 11 OSOLO WAY, AJAO ESTATE LAGOS",
    "MEDPLUS 29 ISAAC JOHN STREET, GRA, IKEJA LAGOS",
    "44 OLU OBASANJO ROAD, ELECHI, PORTHARCOURT, RIVERS STATE (DOMINO'S PIZZA)",
    "MEDPLUS BANEX MALL, PLOT 1, AKIOGUN RD, ONIRU LEKKI, OPPOSITE MAROKO POLICE STATION, LAGOS",
    "642E AKIN ADESOLA STREET VICTORIA ISLAND LAGOS",
    "MEDPLUS STUDIO FLAGSHIP PETER ODILI ROAD GRA, PORT HARCOURT",
    "3, EGBEDA-IDIMU ROAD, DIVINE COURT PLAZA ALIMOSHO BUS STOP, EGBEDA LAGOS",
    "MEDPLUS MAYFAIR GARDEN AWOYAYA LEKKI-EPE EXPRESSWAY LAGOS",
    "MEDPLUS SWEET SENSATION BUILDING, AWOLOWO ROAD IKOYI, LAGOS STATE",
    "MEDPLUS 11, AGUNGI AJIRAN ROAD (SHEPARD PLACE) AGUNGI, LAGOS.",
    "MEDPLUS RANBROOK SQUARE, NO 2 BAALE STREET, IGBO-EFON B/STOP, LEKKI, LAGOS.",
    "MEDPLUS THE PALMS, OTA FORMER GATEWAY HOTEL PREMISES SANGO OTTA",
    "MEDPLUS 41 FREEDOM WAY, LEKKI PHASE 1 LAGOS",
    "MEDPLUS 2 ADEOLA ODEKU STREET, OPP UNION BANK DEJO, VICTORIA ISLAND LAGOS",
    "MEDPLUS 1 AZIKIWE ROAD PORTHARCOURT",
    "GARDEN CITY MALL, 4 ABA ROAD, RUMUOMASI PORT HARCOURT RIVERS STATE",
    "MEDPLUS APAPA MALL, 13 PARK LANE APAPA LAGOS",
    "SHOP 18, SHOPRITE MALL ABIA",
    "MEDPLUS PLOT 310 GBAGADA OWOROSHOKI EXPRESSWAY, GBAGADA LAGOS",
    "MEDPLUS CUBANA WORLD, 17 ADEOLA ODEKU STREET, VI",
    "MEDPLUS 10 CHARITY ROAD, NEW OKO-OBA ABULE EGBA LAGOS",
    "MEDPLUS 31 AWOLOWO ROAD, IKOYI LAGOS STATE",
    "7 ADMIRALTY WAY, LEKKI, LAGOS",
    "MEDPLUS 107 ADENIYI JONES STREET, IKEJA",
    "MEDPLUS DEPARTURE HALL MMA2 IKEJA LAGOS",
    "MEDPLUS 8A KINGSWAY RD, IKOYI LAGOS",
    "MEDPLUS 350/360 IKORODU ROAD MARYLAND, LAGOS.",
    "MEDPLUS BLACKBELL MALL, BESIDE MEGA CHICKEN, IKOTA, LAGOS.",
    "MEDPLUS LANDMARK EVENT CENTRE OPPOSITE HARDROCK CAFÉ WATER CORPORATION DRIVE, ONIRU VICTORIA ISLAND",
    "MEDPLUS 2ND FLOOR, 12 IDOWU MARTINS, VICTORIA ISLAND, LAGOS",
    "MEDPLUS UPTOWN MALL, BARNAWA ROAD BARNAWA PHASE 1, KADUNA STATE",
    "MEDPLUS 113, OGUNLANA DRIVE SURULERE LAGOS",
    "MEDPLUS FOODCO JERICHO - BY JERICHO NURSING HOME, ALL SAINTS ROAD, JERICHO IBADAN",
    "MEDPLUS 1 MUHAMMED STREET, SANTOS LAYOUT, AKOWONJO ROUNDABOUT IKEJA LAGOS",
    "MEDPLUS EVERYDAY SUPERMARKET SUMMIT ROAD CENTRAL AREA ASABA DELTA STATE",
    "MEDPLUS FOODCO RINGROAD - BESIDE IBEDC, RINGROAD IBADAN",
    "PROVIDENCE PLAZA, PLOT 17 & 18 OLOKONLA, SANGOTEDO LAGOS",
    "AHIABA ABAYI OSISIOMA ALONG ABA PORTHARCOURT EXPRESSWAY NGWA LOCAL GOVT AREA ABIA STATE",
    "MEDPLUS CASTLE PRIDE, 226 MURI OKUNOLA, VICTORIA ISLAND LAGOS",
    "MEDPLUS ATLANTIC CENTRE 6 CHEVRON DRIVE, LEKKI.",
    "MEDPLUS TROPICANA MALL, UDO ODOMA AVENUE, UYO AKWA IBOM STATE",
    "LANDMARK EVENT CENTRE OPPOSITE HARDROCK CAFÉ WATER CORPORATION DRIVE, ONIRU VICTORIA ISLAND",
    "MEDPLUS TICKETING AREA, MMA2 ANNEX",
    "MEDPLUS 14 ADEBAYO DOHERTY RD, LEKKI PHASE 1, LAGOS",
    "MEDPLUS 34A FOLA OSIBO STREET, LEKKI",
    "MEDPLUS ORCHID ROAD, AFTER CHEVRON SECOND TOLL GATE, PRIME MALL ENTERTAINMENT LAGOS",
    "MEDPLUS HYPERCITY SUPERMARKET 27 AWOKE STREET, NKPOGU, PORT-HARCOURT",
    "MEDPLUS CITY PLAZA 6 ABAKILIKI STREET AWKA ANAMBRA",
    "MEDPLUS 271B AJOSE ADEOGUN STREET, VICTORIA ISLAND LAGOS",
    "MEDPLUS ONIRU- BLOCK XXI, PLOT 19, CHIEF YESUFU ABIODUN WAY/LAWANI ODULOYE RD, ONIRU ESTATE, LEKKI",
    "MEDPLUS IKEJA MALL, ALAUSA, IKEJA, LAGOS",
    "MEDPLUS 319 RABIU BABATUNDE TINUBU RD, AMUWO ODOFIN.",
    "MEDPLUS 1, AGUIYI IRONSI STREET, MAITAMA ABUJA.",
    "NO 2 PURPLE WAY OFF FREEDOM WAY LEKKI",
    "ASABA SHOPRITE MALL, ASABA",
    "MEDPLUS GAT TERMINAL NNAMDI AZIKIWE INTERNATIONAL AIRPORT, ABUJA.",
    "MEDPLUS FOURTEEN 36 MALL, SANUSI FAFUNWA V.I, LAGOS.",
    "MEDPLUS 11, SHEPERDS PLACE, LEKK, LAGOS.",
    "MEDPLUS THE PALMS, RINGROAD IBADAN OYO STATE IBADAN NIGERIA",
    "NOVARE MALL LEKKI EXPRESSWAY SANGOTEDO LAGOS",
    "NO 3 ILUPEJU BY-PASS ILUPEJU LAGOS",
    "MEDPLUS BNB MALL EMMANUEL LAKOWE GOLF, IBEJU LEKKI, LAGOS.",
    "KWARA MALL, ILORIN KWARA STATE ILORIN",
    "MEDPLUS 10A, ADMIRALTY WAY, LEKKI PHASE 1, OPPOSITE EVER CARE HOSPITAL, LEKKI LAGOS",
    "BRITISH INTERNATIONAL SCHOOL ROAD, LEKKI, LAGOS",
    "MEDPLUS SIMBIAT IKEJA MALL, 22, SIMBIAT ABIOLA WAY, IKEJA LAGOS",
    "MEDPLUS RICHBAM BUILDING, SOKA AREA, NEW FELELE, IBADAN",
    "MEDPLUS ADETOKUNBO ADEMOLA STREET, VICTORIA ISLAND, LAGOS",
    "MEDPLUS BLOCK 113, PLOT 1C HAKEEM DICKSON LEKKI PHASE 1, LEKKI LAGOS",
    "MEDPLUS GARDEN CITY MALL, 4 OLD ABA ROAD, RUMUOMASI PORT HARCOURT RIVERS STATE",
    "MEDPLUS 7 ADMIRALTY WAY LEEKI PHASE I LEKKI LAGOS",
]

STOP = set("""THE A AN AND OR FOR TO IN ON AT BY WITH OF IS IT AS BE THIS THAT FROM
              SHOP SHOPS MEDPLUS STREET ROAD STATE LAGOS NIGERIA""".split())
def norm(s):
    s = re.sub(r"[^A-Za-z0-9 ]", " ", str(s).upper())
    s = re.sub(r"\bMOB0LAJI\b", "MOBOLAJI", s)
    s = re.sub(r"\bPORTHARCOURT\b|\bPORT HARCOURT\b", "PORTHARCOURT", s)
    s = re.sub(r"\bLEKK\b", "LEKKI", s)
    s = re.sub(r"\bSHEPERDS?\b", "SHEPHERDS", s)
    s = re.sub(r"\bSHOPING\b", "SHOPPING", s)
    return [t for t in s.split() if t not in STOP and len(t) > 1]

def jaccard(a, b):
    if not a or not b: return 0.0
    sa, sb = set(a), set(b)
    return len(sa & sb) / len(sa | sb)

# ── Load Medplus.xlsx ──────────────────────────────────────────────────
med_rows = []
raw = pd.read_excel('data/Medplus.xlsx', dtype=str, keep_default_na=False, header=0)
for _, r in raw.iterrows():
    store = str(r.get('STORE NAME', '')).strip()
    med_rows.append({
        'store': store,
        'addr': str(r.get('ADDRESS', '')).strip(),
        'tid': str(r.get('TERMINAL ID', '')).strip(),
        'static': str(r.get('STATIC ACCOUNT', '')).strip(),
        'beneficiary': str(r.get('BENEFICIARY NAME', '')).strip(),
    })

def best_match(tokens):
    best, best_s = None, 0.0
    for row in med_rows:
        for k in ('addr', 'store'):
            rtoks = norm(row[k])
            s = 0.65 * jaccard(tokens, rtoks) + 0.35 * token_sort_ratio(
                ' '.join(tokens), ' '.join(rtoks))
            if s > best_s:
                best_s, best = s, row
    return best, best_s

# ── Verified overrides: fuzzy got these wrong, xlsx cross-check settled them ──
# keyed by substring of the pasted address → (tid, store)
OVERRIDES = {
    'KEFFI STREET, IKOYI': ('2ISW2817', 'MEDPLUS KEFFI'),
    'ASABA SHOPRITE MALL': ('2ISW2564', 'MEDPLUS ASABA'),
    'DELTA MALL EFFURUN': ('2ISW2565', 'MEDPLUS DELTA'),
    'LACIUDAD MALL BLOCK XXVI': ('2ISW2580', 'MEDPLUS LA CUIDAD'),
    'NOVARE MALL LEKKI EXPRESSWAY SANGOTEDO': ('2ISW1920', 'MEDPLUS SANGOTEDO'),
    'ENUGU RETAIL CENTRE LTD': ('2ISW2571', 'MEDPLUS ENUGU'),
    'MMA2 ANNEX': ('2ISW2830', 'MEDPLUS MMA2ANEX'),
}
# Addresses verified NOT present anywhere in Medplus.xlsx
NOT_IN_XLSX = ['OASIS CENTER', 'MARINA LAGOS ISLAND', 'PROVIDENCE PLAZA',
               'BRITISH INTERNATIONAL SCHOOL ROAD', 'FRESHFORTE',
               '39 LEKKI ESTATE ROAD WITHIN BELA VISTA']

by_tid = {r['tid']: r for r in med_rows}

results = []
for i, addr in enumerate(ADDRESSES, 1):
    up = addr.upper()
    ov = next(((k, v) for k, v in OVERRIDES.items() if k in up), None)
    if any(n in up for n in NOT_IN_XLSX):
        results.append({'no': i, 'address': addr, 'status': 'NOT FOUND',
                        'store': '', 'tid': '', 'static': '', 'beneficiary': '', 'score': 0.0})
        continue
    if ov:
        tid, store = ov[1]
        row = by_tid.get(tid, {})
        results.append({'no': i, 'address': addr, 'status': 'MATCHED',
                        'store': store, 'tid': tid,
                        'static': row.get('static', ''), 'beneficiary': row.get('beneficiary', ''),
                        'score': 0.99})
        continue
    toks = norm(addr)
    m, s = best_match(toks)
    status = 'MATCHED'
    if not m or s < 0.50:
        status = 'LOW CONFIDENCE' if (m and s >= 0.35) else 'NOT FOUND'
    results.append({
        'no': i, 'address': addr, 'status': status,
        'store': m['store'] if m else '', 'tid': m['tid'] if m else '',
        'static': m['static'] if m else '', 'beneficiary': m['beneficiary'] if m else '',
        'score': round(s, 3) if m else 0.0,
    })

matched = [r for r in results if r['status'] == 'MATCHED']
low = [r for r in results if r['status'] == 'LOW CONFIDENCE']
nf = [r for r in results if r['status'] == 'NOT FOUND']
print(f"matched={len(matched)}  low-confidence={len(low)}  not-found={len(nf)}  total={len(results)}")
print("\n-- NOT FOUND --")
for r in nf:
    print(f"  #{r['no']:>3}  {r['address'][:75]}")
print("\n-- LOW CONFIDENCE (review manually) --")
for r in sorted(low, key=lambda x: x['score']):
    print(f"  #{r['no']:>3} {r['score']:.2f}  {r['tid']:<10} {r['store'][:32]:<32} :: {r['address'][:55]}")
print("\n-- LOWEST 6 MATCHED --")
for r in sorted(matched, key=lambda x: x['score'])[:6]:
    print(f"  #{r['no']:>3} {r['score']:.2f}  {r['tid']:<10} {r['store'][:32]:<32} :: {r['address'][:55]}")

# ── Excel export ───────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Medplus TIDs"

headers = ['#', 'Status', 'Store Name', 'TID', 'Static Account', 'Beneficiary',
           'Pasted Address', 'Match Score']
ws.append(headers)

hdr_fill = PatternFill('solid', fgColor='1F4E79')
hdr_font = Font(bold=True, color='FFFFFF', size=11)
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

green = PatternFill('solid', fgColor='E2EFDA')
yellow = PatternFill('solid', fgColor='FFF2CC')
red = PatternFill('solid', fgColor='FCE4EC')

for r in results:
    ws.append([r['no'], r['status'], r['store'], r['tid'], r['static'], r['beneficiary'],
               r['address'], r['score']])
    rr = ws.max_row
    fill = green if r['status'] == 'MATCHED' else (yellow if r['status'] == 'LOW CONFIDENCE' else red)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=rr, column=c)
        cell.border = border
        if c == 2:
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
        if c == 8:
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal='center')

widths = [5, 15, 32, 11, 15, 30, 62, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:H{ws.max_row}"

out = Path('data/medplus_tids.xlsx')
wb.save(out)
print(f"\nSaved: {out}  ({len(results)} rows)")
