import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Merchants Contact List"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
nf = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ff = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
tb = Border(left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))

for col, h in enumerate(["Merchant Name", "Contact Name(s)", "Phone Number(s)", "Email(s)", "Status"], 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")
    c.border = tb

data = [
    ("ABC CARGO EXPRESS LIMITED 1", "EMMA NJOKU", "+2348120452526", "emmanjoku@abctransport.com", "Found"),
    ("ABC CARGO EXPRESS LTD (BENIN)", "ABASIAMA ABASIEKONG", "7061517637", "olivianwogu2004@gmail.com", "Found"),
    ("ABC CARGO EXPRESS LTD (CALABAR)", "ETIM IMOH OKON", "9091902265", "olivianwogu2004@gmail.com", "Found"),
    ("ABC CARGO EXPRESS LTD (FCT ABUJA)", "NWOKORIE CHINEDU", "8037966161", "olivianwogu2004@gmail.com", "Found"),
    ("ABC CARGO EXPRESS LTD (KADUNA)", "AREMU OLUWAGBEMIGA", "8066287966", "olivianwogu2004@gmail.com", "Found"),
    ("ABC CARGO EXPRESS LTD (PORTHARCOURT)", "UGWUADU UDOCHUKWU CYRIACUS", "7067692599", "olivianwogu2004@gmail.com", "Found"),
    ("ABC CARGO EXPRESS LTD (-OJO)", "EKWEOGU CHIMAOBI", "7030675372", "olivianwogu2004@gmail.com", "Found"),
    ("ABC CARGO EXPRESS LTD (ONITSHA)", "CHINKATA CHIMEZIE", "8189205958", "olivianwogu2004@gmail.com", "Found"),
    ("ABC CARGO EXPRESS LTD (AMUWO)", "OLIVIA NWOGU", "08034100549", "olivianwogu2004@gmail.com", "Found"),
    ("ABC CARGO EXPRESS LTD (BOLADE LAGOS)", "IBEH GENEVIEVE AKARACHI", "08033672339", "genevieveibeh@yahoo.com", "Found"),
    ("ABC CARGO EXPRESS LTD (JIBOWU)", "SANNI ANUOLUWA", "08110237902", "sannidamilola74@gmail.com", "Found"),
    ("ABC CARGO EXPRESS LTD (OSHODI)", "CHINWENDU OHAERI", "08059623016", "ohaerichinwendu69@gmail.com", "Found"),
    ("ABC CARGO EXPRESS LTD (PORTHARCOURT 2)", "AMECHI VINCENT UDO.", "08037517174", "Vincent_amechi@yahoo.com", "Found"),
    ("ABC CARGO EXPRESS LTD (UTAKO)", "MBA CHIBUIKE MODESTUS", "08066703801", "Mbahchibuike122@gmail", "Found"),
    ("ADDIDE ABARANJE", "ADE ODUNUGA", "8125223370", "B37abaranje@addide.com", "Found"),
    ("ADDIDE AGBADO", "ADE ODUNUGA", "9078542730", "C42agbado@addide.com", "Found"),
    ("ADDIDE AGUDA", "ADE ODUNUGA", "8086725904", "B25aguda@addide.com", "Found"),
    ("ADDIDE AJAYI", "ADE ODUNUGA", "8086200529", "A13ajayi@addide.com", "Found"),
    ("ADDIDE AJUWON", "ADE ODUNUGA", "7015423465", "B39ajuwon@addide.com", "Found"),
    ("ADDIDE AKOKA", "ADE ODUNUGA", "8024335018", "a02akoka@addide.com", "Found"),
    ("ADDIDE AKUTE", "ADE ODUNUGA", "8086725829", "B22akute@addide.com", "Found"),
    ("ADDIDE ALAPERE", "ADE ODUNUGA", "8024258454", "B31alapere@addide.com", "Found"),
    ("ADDIDE APATA", "ADE ODUNUGA", "8024335032", "A06apata@addide.com", "Found"),
    ("ADDIDE AROWOJOBE", "ADE ODUNUGA", "8020928073", "B33arowojobe@addide.com", "Found"),
    ("ADDIDE BAJULAYE", "ADE ODUNUGA", "7014010880", "c47bajulaye@addide.com", "Found"),
    ("ADDIDE BARUWA", "ADE ODUNUGA", "8123140697", "c48baruwa@addide.com", "Found"),
    ("ADDIDE CHARITY", "ADE ODUNUGA", "70146770103", "B26okooba@addide.com", "Found"),
    ("ADDIDE COMMAND", "ADE ODUNUGA", "8086200437", "A12command@addide.com", "Found"),
    ("ADDIDE DEMURIN", "ADE ODUNUGA", "8024335067", "A07demurin@addide.com", "Found"),
    ("ADDIDE DIYA", "ADE ODUNUGA", "8024335021", "A03diya@addide.com", "Found"),
    ("ADDIDE DOPEMU", "ADE ODUNUGA", "7082447592", "A10dopemu@addide.com", "Found"),
    ("ADDIDE EBUTE", "ADE ODUNUGA", "8086725955", "B23ebute@addide.com", "Found"),
    ("ADDIDE EGBE", "ADE ODUNUGA", "8027658327", "B32egbe@addide.com", "Found"),
    ("ADDIDE FAGBA", "ADE ODUNUGA", "7084041565", "B28fagba@addide.com", "Found"),
    ("ADDIDE GOV RD", "ADE ODUNUGA", "8127666601", "B36governorroad@addide.com", "Found"),
    ("ADDIDE IDIMU", "ADE ODUNUGA", "8081726506", "B34idimu@addide.com", "Found"),
    ("ADDIDE IGBOGBO", "ADE ODUNUGA", "9078011921", "C41igbogbo@addide.com", "Found"),
    ("ADDIDE IJEDE", "ADE ODUNUGA", "7082345146", "C44ijede@addide.com", "Found"),
    ("ADDIDE IJEGUN", "ADE ODUNUGA", "7013870522", "B38ijegun@addide.com", "Found"),
    ("ADDIDE IJESHA", "ADE ODUNUGA", "8024335028", "A05ijesha@addide.com", "Found"),
    ("ADDIDE IJU", "ADE ODUNUGA", "8024780908", "A16iju@addide.com", "Found"),
    ("ADDIDE IKORODU", "ADE ODUNUGA", "8024780900", "A18benson@addide.com", "Found"),
    ("ADDIDE IKOSI", "ADE ODUNUGA", "8086658073", "A15ikosi@addide.com", "Found"),
    ("ADDIDE ISAWO", "ADE ODUNUGA", "8086725830", "B21isawo@addide.com", "Found"),
    ("ADDIDE ITIRE", "ADE ODUNUGA", "8086725895", "B24itire@addide.com", "Found"),
    ("ADDIDE IWAYA", "ADE ODUNUGA", "8024335079", "A08iwaya@addide.com", "Found"),
    ("ADDIDE LADIPO", "ADE ODUNUGA", "8024780909", "A19ladipo@addide.com", "Found"),
    ("ADDIDE LAISU", "ADE ODUNUGA", "9079185044", "C43laisu@addide.com", "Found"),
    ("ADDIDE LIMITED (HQ)", "ADE ODUNUGA", "8024780894", "A11ogba@addide.com", "Found"),
    ("ADDIDE MAFOLUKU", "ADE ODUNUGA", "7086647898", "B29mafoluku@addide.com", "Found"),
    ("ADDIDE MORGAN", "ADE ODUNUGA", "7087983066", "B30morgan@addide.com", "Found"),
    ("ADDIDE ODOGUYAN", "ADE ODUNUGA", "7083474760", "c46odongunyan@addide.com", "Found"),
    ("ADDIDE OGBA", "ADE ODUNUGA", "7082447447", "A11ogba@addide.com", "Found"),
    ("ADDIDE OJODU", "ADE ODUNUGA", "8086657647", "A14ojodu@addide.com", "Found"),
    ("ADDIDE PEDRO", "ADE ODUNUGA", "8024335025", "A04pedro@addide.com", "Found"),
    ("ADDIDE POWERLINE", "ADE ODUNUGA", "8024780894", "A17powerline@addide.com", "Found"),
    ("ADDIDE SANTO", "ADE ODUNUGA", "8024780894", "c49abuleegba@addide.com", "Found"),
    ("ADDIDEOLD OTA", "ADE ODUNUGA", "7083468662", "B35oldotaroad@addide.com", "Found"),
    ("ADENIKE AGORO", "ADENIKE AGORO", "2348139939829", "wolemarconcept@gmail.com", "Found"),
    ("A-PURE LIFESTYLE PHARM", "MR (no name in DB)", "+2348165792450", "m.ajayi@purelifepharmacy.ng", "Found"),
    ("ARINZE VETFOODS", "ARINZE OGIDI", "08068046910", "vetwealth@yahoo.com", "Found"),
    ("ARTEE INDUSTRIES LIMITED", "DAVID / SURESH", "08129925099 / 08129925096",
     "account.treasury@arteegroup.com, suresh.mk@arteegroup.com", "Found"),
    ("ATREOS RETAIL PLATFORM LIMITED", "OLUWATOYIN OSHO", "09111025393", "NBASHIR@ATREOS.COM", "Found"),
    ("ATREOS RETAIL PLATFORM-AJEGUNLE", "OLUWATOYIN OSHO", "09111025393", "NBASHIR@ATREOS.COM", "Found"),
    ("ATREOS RETAIL PLATFORM-FADEYI", "OLUWATOYIN OSHO", "09111025393", "NBASHIR@ATREOS.COM", "Found"),
    ("ATREOS RETAIL PLATFORM-FAKOREDE", "OLUWATOYIN OSHO", "09111025393", "NBASHIR@ATREOS.COM", "Found"),
    ("AZZ LIMITED (PTSP KEYSTONE)", "-", "-", "-", "Not Found"),
    ("BEACONHEALTH (SANGOTEDO)", "RUTH AKINWALE", "7068430600", "ruth@mdaasnigeria.com", "Found"),
    ("BEACON HEALTH BODIJA", "OLUWAFEMI ADEOSUN", "08154097568", "bodija.manager@beaconhealth.io", "Found"),
    ("BEACON HEALTH HAKEEM DICKSON", "BAKARE QUDUS FOLORUNSHO", "09053444625", "hd@beaconhealth.io", "Found"),
    ("BEACON HEALTH ADO", "NIYI ALUKO", "08036433009", "adomanager@beaconhealth.io", "Found"),
    ("BEACON HEALTH PH", "AMOS EMMANUEL", "08150914850", "ph.manager@beaconhealth.io", "Found"),
    ("BEACON HEALTH BENIN", "ELIAS MOHIE", "08057304171", "benin.manager@beaconhealth.io", "Found"),
    ("CARELINK INTEGRATED SERVICES LTD", "OLORIEGBE HAFSAT ONJIMOH", "08037082318",
     "mhafsat2000@yahoo.com", "Found"),
    ("CASCADES LUXURY LIMITED", "-", "-", "-", "Not Found"),
    ("DGENNY GLOBAL LINK LIMITED", "ABIOLA FABOYEDE", "08023155901",
     "dgennygloballinks@gmail.com", "Found"),
    ("EAGLE FLIGHT MICROFINANCE BANK LTD", "W.L.B.C. EAGLE STRENGTH BOOKSHOP", "+2348038249913",
     "rehobothinterntionalschools@yahoo.co", "Found"),
    ("EMUREN JOSEPHINE IFE-ERE", "JOSEPHINE EMUREN", "9155555642", "emurenjay@gmail.com", "Found"),
    ("FEMI AYODELE", "FEMI AYODELE", "8073230000", "wakkisfood@gmail.com", "Found"),
    ("FOLASHADE KALEJAIYE", "FOLASHADE KALEJAIYE", "+2348082132474", "phola.isaac@gmail.com",
     "Found"),
    ("FOREVER LIVING PRODUCTS NIG LTD", "ADEWALE OGUNDARE", "08026650654",
     "warri@flpng.com, victoriaisland@flpng.com, enugu@flpng.com, port-harcourt@flpng.com, "
     "benin-city@flpng.com, kano@flpng.com, abuja@flpng.com, asaba@flpng.com", "Found"),
    ("FRANKLIN IBEDU", "FRANKLIN IBEDU", "08087610776", "okeytufan2@gmail.com", "Found"),
    ("GARKUWA SPECIALIST HOSPITAL", "-", "-", "-", "Not Found"),
    ("GARTMORE INVESTMENTS LIMITED", "Trevor Pillay", "8066483473", "trevor@kadacinemas.com",
     "Found"),
    ("GLADYS ARARAUME", "GLADYS ARARAUME", "+2348162623277",
     "exclusivestoreslimited1202@gmail.com", "Found"),
    ("ISIKILU RAJI SEGUN", "ISIKILU RAJI SEGUN", "+2348137239599", "Aremuraji1@yahoo.com",
     "Found"),
    ("JOY EFI", "-", "-", "-", "Not Found"),
    ("JULIANA BOT", "MR (JJ STORE contact)", "8064106801", "winsfree@mail.com", "Found"),
    ("JULIES LAUNJA", "CHINEDU ANYANWU (MR)", "08033273875",
     "chinedu.anyanwu1973@gmail.com", "Found"),
    ("LA BELLE BISTRO AND BAR", "-", "-", "-", "Not Found"),
    ("MARGRET STORE", "ONAH OGECHI MARGRET (MRS)", "7033938311 / 8039443711",
     "Onahogechimargret@gmail.com", "Found"),
    ("MAX AIR", "KABIR YUSUF BAKORI", "08037737787", "kybakori@yahoo.com", "Found"),
    ("MIMSHACK WHOLESALE SHOP", "OLASUMBO OLIJOGUN BEATRICE", "8132381580",
     "Oolijogun@gmail.com", "Found"),
    ("MJ RESTAURANT", "JOY MICHAEL", "08069713344", "MJRestaurant2022@gmail.com", "Found"),
    ("MOs Food Mart", "NWASOKA UGOCHUKWU", "08034158970", "maureennwasoka@gmail.com", "Found"),
    ("NEST OGB LIMITED_FIDELITY", "-", "-", "-", "Not Found"),
    ("NEWHEALTH PHARMACY LTD 3", "-", "-", "-", "Not Found"),
    ("NIGERIA POLICE FORCE MFB", "-", "-", "-", "Not Found"),
    ("NUCLEAR CLEAR-BIZ & ENTERPRISE", "-", "-", "-", "Not Found"),
    ("ODICHUKWU CHIGOZIE CHARLES", "CHIGOZIE ODICHUKWU", "08060304226",
     "odichukwucharles@gmail.com", "Found"),
    ("OFFICE R US LIMITED", "OLUFEMI PETER / LAKSH", "09098882522 / 09098882801",
     "accounts@office-r-us.com, Laksh@office-r-us.com", "Found"),
    ("OGEO GLOBAL INVESTMENT NIG LTD", "-", "-", "-", "Not Found"),
    ("OLISON SUPER STORES", "-", "-", "-", "Not Found"),
    ("ONYEKA STORES", "NNAJI UNAMUOKA ONYEKA", "07069239412",
     "ogagajessica354@gmail.com", "Found"),
    ("PHIL HALLMARK SUPERMARKET", "OSAMUDIAMEN EDIGIN", "8109419498",
     "mudi147@yahoo.com", "Found"),
    ("PIK_WIK NIGERIA", "-", "-", "-", "Not Found"),
    ("PINANCLE SUPERMARKET", "-", "-", "-", "Not Found"),
    ("PROCEPT EYE CLINIC", "OBEHI MARGARET OSOATA", "8176512585",
     "procepteyeclinic@gmail.com", "Found"),
    ("REIZ CONTINENTAL HOTELS LTD",
     "ADEBISI JELILI ADEBAYO / CASMIR ANOSIKE AKUCHIE",
     "7032057996 / 8037026531",
     "reizcontinentalhotelabuja@gmail.com, info@reizcontinentalhotels.com", "Found"),
    ("ROBAN STORES NIG. LTD", "-", "-", "-", "Not Found"),
    ("ROSYFRANK GLOBAL VENTURES", "-", "-", "-", "Not Found"),
    ("TINA OKI", "TINA OKI", "08039689799", "Okidoris77@gmail.com", "Found"),
    ("WORTHWELL PHARMACEUTICALS COMPANY LTD", "CHYLOTTEM NDIUKWU", "08037556085",
     "juoffny@yahoo.com, worthwellpharmacy@gmail.com", "Found"),
]

for ri, (m, co, ph, em, st) in enumerate(data, 2):
    ws.cell(row=ri, column=1, value=m).border = tb
    ws.cell(row=ri, column=2, value=co).border = tb
    ws.cell(row=ri, column=3, value=ph).border = tb
    ws.cell(row=ri, column=4, value=em).border = tb
    c5 = ws.cell(row=ri, column=5, value=st)
    c5.border = tb
    if st == "Found":
        c5.fill = ff
        c5.font = Font(color="006100")
    else:
        c5.fill = nf
        c5.font = Font(color="9C5700")

ws.column_dimensions["A"].width = 55
ws.column_dimensions["B"].width = 48
ws.column_dimensions["C"].width = 35
ws.column_dimensions["D"].width = 75
ws.column_dimensions["E"].width = 12
ws.auto_filter.ref = "A1:E%d" % (len(data) + 1)
ws.freeze_panes = "A2"
wb.save("Merchants_Contact_List.xlsx")
found = sum(1 for d in data if d[4] == "Found")
notf = sum(1 for d in data if d[4] != "Found")
print("Saved: Merchants_Contact_List.xlsx (%d rows) - Found: %d | Not Found: %d" % (len(data), found, notf))
