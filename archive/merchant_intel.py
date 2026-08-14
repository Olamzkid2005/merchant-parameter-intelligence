"""
Merchant Intelligence Search Tool
==================================
Phases 1-10: Intelligent workbook discovery, alias engine, token intelligence,
row search, field scoring, email discovery, similar merchants, duplicate detection,
merchant intelligence report, and automatic learning.

SQLite FTS5 backend for instant searches on future runs.
"""

import os
import re
import json
import sqlite3
import hashlib
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from difflib import SequenceMatcher

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rapidfuzz import fuzz, process

# ─── Configuration ───────────────────────────────────────────────────────────

WORKBOOK = r"C:\Users\David.Olamijulo\downloads\parameter\data\2ISW_Parameter_File 5.xlsx"
DB_PATH = r"C:\Users\David.Olamijulo\Downloads\parameter\merchant_intel.db"
REPORT_PATH = r"C:\Users\David.Olamijulo\Downloads\parameter\Merchant_Report.xlsx"
LEARNING_FILE = r"C:\Users\David.Olamijulo\downloads\parameter\data\merchant_aliases.json"

MERCHANT_COL_KEYWORDS = [
    "merchant name", "merchant_name", "merchantname", "business name", "trading name",
    "outlet", "slip header", "slip_header", "slipheader", "account name",
    "account_name", "settlement name", "legal name", "company name", "store name",
    "merchant", "name",
]

EMAIL_COL_KEYWORDS = ["email", "e-mail", "mail", "email alerts", "email_alerts"]
PHONE_COL_KEYWORDS = ["mobile phone", "mobilephone", "mobile_phone", "phone", "telephone", "mobile", "phone no", "phone number"]
CONTACT_COL_KEYWORDS = ["contact name", "contact_name", "contactname", "contact", "contact person"]
ADDRESS_COL_KEYWORDS = ["physical addr", "physicaladdr", "physical_addr", "address", "merchantphysicaladdr", "merchant address", "terminal address"]
TID_COL_KEYWORDS = ["terminal id", "terminal_id", "terminalid", "tid"]
MERCHANT_ID_KEYWORDS = ["merchant id", "merchant_id", "merchantid"]
PTSP_KEYWORDS = ["ptsp", "ptspcode", "ptsp_code"]
STATE_KEYWORDS = ["state code", "state_code", "statecode"]
BANK_KEYWORDS = ["bank code", "bankcode", "bank_code", "bank"]
SLIP_HEADER_KEYWORDS = ["slip header", "slip_header", "slipheader"]
CONTACT_TITLE_KEYWORDS = ["contact title", "contacttitle", "contact_title"]

GENERIC_WORDS = [
    "LTD", "LIMITED", "NIGERIA", "NIG", "GLOBAL", "SERVICES", "ENTERPRISES",
    "ENTERPRISE", "INVESTMENT", "INVESTMENTS", "PLC", "CORPORATION", "CORP",
    "GROUP", "HOLDINGS", "HOLDING", "SOLUTIONS", "TECHNOLOGIES", "TECHNOLOGY",
    "VENTURES", "CONCEPTS", "RESOURCES", "INTEGRATED", "NETWORK", "NETWORKS",
    "SYSTEMS", "INTERNATIONAL", "INTL", "ASSOCIATES", "PARTNERS", "TRADING",
    "INDUSTRIES", "COMPANY", "CO", "UNLIMITED", "LTD liability company",
    "AGENCY", "AGENCIES", "PROJECTS", "PROJECT", "LOGISTICS", "SUPPLIES",
    "DISTRIBUTION", "DISTRIBUTORS", "NIGERIAN", "AND", "THE", "OF", "FOR",
    "NIG LTD", "NIGERIA LTD",
]

STOP_TOKENS = {"THE", "A", "AN", "OF", "FOR", "AND", "IN", "TO", "AT", "BY", "ON", "IS", "IT", "AS", "BE", "OR", "&"}

# ─── Column Discovery ────────────────────────────────────────────────────────

def discover_columns(df):
    mapping = {}
    for col in df.columns:
        col_clean = str(col).replace("\xa0", "").strip().lower()
        if any(kw in col_clean for kw in MERCHANT_COL_KEYWORDS):
            if "merchant" in col_clean and ("name" in col_clean or "account" not in col_clean):
                mapping.setdefault("merchant_name", col)
        if any(kw in col_clean for kw in EMAIL_COL_KEYWORDS):
            mapping.setdefault("email", col)
        if any(kw in col_clean for kw in PHONE_COL_KEYWORDS):
            mapping.setdefault("mobile_phone", col)
        if any(kw in col_clean for kw in CONTACT_COL_KEYWORDS):
            mapping.setdefault("contact_name", col)
        if any(kw in col_clean for kw in ADDRESS_COL_KEYWORDS):
            mapping.setdefault("physical_addr", col)
        if any(kw in col_clean for kw in TID_COL_KEYWORDS):
            mapping.setdefault("terminal_id", col)
        if any(kw in col_clean for kw in MERCHANT_ID_KEYWORDS):
            mapping.setdefault("merchant_id", col)
        if any(kw in col_clean for kw in PTSP_KEYWORDS):
            mapping.setdefault("ptsp", col)
        if any(kw in col_clean for kw in STATE_KEYWORDS):
            mapping.setdefault("state_code", col)
        if any(kw in col_clean for kw in BANK_KEYWORDS):
            mapping.setdefault("bank_code", col)
        if any(kw in col_clean for kw in SLIP_HEADER_KEYWORDS):
            mapping.setdefault("slip_header", col)
        if any(kw in col_clean for kw in CONTACT_TITLE_KEYWORDS):
            mapping.setdefault("contact_title", col)
        if "account" in col_clean and "name" in col_clean and "merchant" in col_clean:
            mapping.setdefault("account_name", col)
    return mapping


def clean_val(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    s = s.replace("\xa0", " ").replace("\ufffd", "")
    return s


# ─── Alias Engine ────────────────────────────────────────────────────────────

def generate_aliases(name):
    if not name:
        return []
    name_upper = name.upper().strip()
    aliases = set()
    aliases.add(name.strip())
    aliases.add(name_upper)

    tokens = name_upper.split()
    filtered = [t for t in tokens if t not in GENERIC_WORDS and len(t) > 1]
    if filtered:
        aliases.add(" ".join(filtered))
        for t in filtered:
            aliases.add(t)

    for _ in range(3):
        new_tokens = [t for t in tokens if t not in GENERIC_WORDS]
        if len(new_tokens) < len(tokens):
            tokens = new_tokens
            if tokens:
                aliases.add(" ".join(tokens))
                if len(tokens) == 1 and len(tokens[0]) > 1:
                    aliases.add(tokens[0])
        else:
            break

    for kw in GENERIC_WORDS:
        pattern = re.compile(r"\s*" + re.escape(kw) + r"\s*", re.IGNORECASE)
        stripped = pattern.sub(" ", name).strip()
        if stripped and stripped != name.strip():
            aliases.add(stripped)
            aliases.add(stripped.upper())

    name_no_special = re.sub(r"[^A-Z0-9\s]", " ", name_upper).strip()
    name_no_special = re.sub(r"\s+", " ", name_no_special)
    if name_no_special != name_upper:
        aliases.add(name_no_special)
        filtered_ns = [t for t in name_no_special.split() if t not in GENERIC_WORDS and len(t) > 1]
        if filtered_ns:
            aliases.add(" ".join(filtered_ns))

    return sorted(a for a in aliases if len(a) > 1)


# ─── Token Intelligence ──────────────────────────────────────────────────────

def extract_significant_tokens(name):
    if not name:
        return []
    tokens = name.upper().split()
    return [t for t in tokens if t not in GENERIC_WORDS and t not in STOP_TOKENS and len(t) > 1]


def tokenize_for_search(text):
    if not text:
        return ""
    text = re.sub(r"[^A-Z0-9\s]", " ", text.upper())
    text = re.sub(r"\s+", " ", text).strip()
    tokens = text.split()
    return " ".join(t for t in tokens if t not in STOP_TOKENS and len(t) > 1)


# ─── Workbook Loader ─────────────────────────────────────────────────────────

class WorkbookLoader:
    def __init__(self, path):
        self.path = path
        self.sheets = {}
        self.records = []
        self._load()

    def _load(self):
        xls = pd.ExcelFile(self.path)
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            df = df.dropna(axis=1, how="all")
            mapping = discover_columns(df)
            self.sheets[sheet_name] = {"df": df, "mapping": mapping}
        self._extract_records()

    def _extract_records(self):
        for sheet_name, info in self.sheets.items():
            df = info["df"]
            mapping = info["mapping"]
            for idx, row in df.iterrows():
                rec = {
                    "sheet_name": sheet_name,
                    "row_num": idx + 2,
                    "merchant_name": "",
                    "slip_header": "",
                    "email": "",
                    "mobile_phone": "",
                    "contact_name": "",
                    "contact_title": "",
                    "physical_addr": "",
                    "account_name": "",
                    "terminal_id": "",
                    "merchant_id": "",
                    "ptsp": "",
                    "state_code": "",
                    "bank_code": "",
                    "raw": {},
                }
                for field_key, col_name in mapping.items():
                    if col_name in row:
                        rec[field_key] = clean_val(row[col_name])
                for col in df.columns:
                    rec["raw"][str(col)] = clean_val(row.get(col, ""))
                self.records.append(rec)

    def get_all_text_for_fts(self, rec):
        fields = [
            rec["merchant_name"], rec["slip_header"], rec["email"],
            rec["mobile_phone"], rec["contact_name"], rec["contact_title"],
            rec["physical_addr"], rec["account_name"], rec["terminal_id"],
            rec["merchant_id"],
        ]
        return " ".join(str(f) for f in fields if f)


# ─── SQLite FTS5 Database ────────────────────────────────────────────────────

class MerchantDB:
    def __init__(self, db_path):
        self.db_path = db_path
        self.conn = sqlite3.connect(db_path)
        self.conn.execute("PRAGMA journal_mode=WAL")
        self._init_schema()

    def _init_schema(self):
        c = self.conn.cursor()
        c.executescript("""
            CREATE TABLE IF NOT EXISTS merchants (
                id INTEGER PRIMARY KEY,
                sheet_name TEXT,
                row_num INTEGER,
                merchant_name TEXT,
                slip_header TEXT,
                email TEXT,
                mobile_phone TEXT,
                contact_name TEXT,
                contact_title TEXT,
                physical_addr TEXT,
                account_name TEXT,
                terminal_id TEXT,
                merchant_id TEXT,
                ptsp TEXT,
                state_code TEXT,
                bank_code TEXT,
                raw_json TEXT
            );
            CREATE VIRTUAL TABLE IF NOT EXISTS merchants_fts USING fts5(
                merchant_name, slip_header, email, mobile_phone,
                contact_name, contact_title, physical_addr,
                account_name, terminal_id, merchant_id,
                tokenize='porter unicode61'
            );
            CREATE TABLE IF NOT EXISTS aliases (
                canonical TEXT,
                alias TEXT,
                source TEXT DEFAULT 'auto',
                UNIQUE(canonical, alias)
            );
            CREATE TABLE IF NOT EXISTS email_index (
                merchant_name TEXT,
                email TEXT,
                sheet_name TEXT,
                row_num INTEGER
            );
            CREATE TABLE IF NOT EXISTS learning_log (
                discovered TEXT,
                canonical TEXT,
                confidence REAL,
                timestamp TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_merchant_name ON merchants(merchant_name);
            CREATE INDEX IF NOT EXISTS idx_alias ON aliases(alias);
        """)
        self.conn.commit()

    def clear_data(self):
        c = self.conn.cursor()
        for table in ["merchants", "merchants_fts", "email_index", "learning_log"]:
            c.execute(f"DELETE FROM {table}")
        self.conn.commit()

    def insert_merchant(self, rec):
        c = self.conn.cursor()
        raw_json = json.dumps(rec["raw"], default=str)
        c.execute("""
            INSERT INTO merchants
                (sheet_name, row_num, merchant_name, slip_header, email,
                 mobile_phone, contact_name, contact_title, physical_addr,
                 account_name, terminal_id, merchant_id, ptsp, state_code,
                 bank_code, raw_json)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            rec["sheet_name"], rec["row_num"], rec["merchant_name"],
            rec["slip_header"], rec["email"], rec["mobile_phone"],
            rec["contact_name"], rec["contact_title"], rec["physical_addr"],
            rec["account_name"], rec["terminal_id"], rec["merchant_id"],
            rec["ptsp"], rec["state_code"], rec["bank_code"], raw_json,
        ))
        merch_id = c.lastrowid
        fts_text = " ".join(str(rec.get(f, "")) for f in [
            "merchant_name", "slip_header", "email", "mobile_phone",
            "contact_name", "contact_title", "physical_addr",
            "account_name", "terminal_id", "merchant_id",
        ])
        cols = "merchant_name, slip_header, email, mobile_phone, contact_name, contact_title, physical_addr, account_name, terminal_id, merchant_id"
        c.execute(f"INSERT INTO merchants_fts(rowid, {cols}) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                  (merch_id, rec.get("merchant_name",""), rec.get("slip_header",""), rec.get("email",""),
                   rec.get("mobile_phone",""), rec.get("contact_name",""), rec.get("contact_title",""),
                   rec.get("physical_addr",""), rec.get("account_name",""), rec.get("terminal_id",""),
                   rec.get("merchant_id","")))
        if rec["email"]:
            c.execute("INSERT INTO email_index VALUES (?,?,?,?)",
                      (rec["merchant_name"], rec["email"], rec["sheet_name"], rec["row_num"]))
        self.conn.commit()
        return merch_id

    def insert_alias(self, canonical, alias, source="auto"):
        c = self.conn.cursor()
        try:
            c.execute("INSERT OR IGNORE INTO aliases VALUES (?,?,?)", (canonical, alias, source))
            self.conn.commit()
        except sqlite3.IntegrityError:
            pass

    def log_learning(self, discovered, canonical, confidence):
        c = self.conn.cursor()
        c.execute("INSERT INTO learning_log VALUES (?,?,?,?)",
                  (discovered, canonical, confidence, datetime.now().isoformat()))
        self.conn.commit()

    def search_fts(self, query, limit=100):
        c = self.conn.cursor()
        query_terms = tokenize_for_search(query)
        if not query_terms:
            return []
        fts_query = " AND ".join(f'"{t}"' for t in query_terms.split())
        try:
            c.execute("""
                SELECT m.*, rank FROM merchants m
                JOIN merchants_fts fts ON m.id = fts.rowid
                WHERE merchants_fts MATCH ?
                ORDER BY rank
                LIMIT ?
            """, (fts_query, limit))
            return c.fetchall()
        except sqlite3.OperationalError:
            return []

    def search_by_alias(self, query):
        c = self.conn.cursor()
        query_upper = query.upper().strip()
        c.execute("SELECT canonical FROM aliases WHERE alias = ?", (query_upper,))
        row = c.fetchone()
        if row:
            canonical = row[0]
            c.execute("""
                SELECT * FROM merchants
                WHERE merchant_name = ? OR merchant_name LIKE ?
            """, (canonical, f"%{canonical}%"))
            return c.fetchall()
        return []

    def get_all_merchant_names(self):
        c = self.conn.cursor()
        c.execute("SELECT DISTINCT merchant_name FROM merchants WHERE merchant_name != '' ORDER BY merchant_name")
        return [r[0] for r in c.fetchall()]

    def get_emails_for_merchant(self, merchant_name):
        c = self.conn.cursor()
        c.execute("""
            SELECT DISTINCT email, sheet_name, row_num FROM email_index
            WHERE merchant_name = ? AND email != ''
            ORDER BY email
        """, (merchant_name,))
        return c.fetchall()

    def close(self):
        self.conn.close()


# ─── Scoring Engine ──────────────────────────────────────────────────────────

def score_field(query, field_value):
    if not query or not field_value:
        return 0
    q = query.upper().strip()
    fv = field_value.upper().strip()
    if q == fv:
        return 100
    if q in fv:
        base = 80
        bonus = min(15, len(q) / len(fv) * 20)
        return int(base + bonus)
    ratio = fuzz.token_sort_ratio(q, fv)
    partial = fuzz.partial_ratio(q, fv)
    score = max(ratio * 0.8 + partial * 0.2, partial * 0.6)
    return int(score)


def score_record(query, rec):
    fields = {
        "Merchant Name": rec.get("merchant_name", ""),
        "Slip Header": rec.get("slip_header", ""),
        "Account Name": rec.get("account_name", ""),
        "Address": rec.get("physical_addr", ""),
        "Email": rec.get("email", ""),
        "Contact": rec.get("contact_name", ""),
        "Contact Title": rec.get("contact_title", ""),
    }
    scores = {}
    for field_name, value in fields.items():
        scores[field_name] = score_field(query, value)
    weighted = (
        scores.get("Merchant Name", 0) * 0.35
        + scores.get("Slip Header", 0) * 0.20
        + scores.get("Account Name", 0) * 0.15
        + scores.get("Email", 0) * 0.10
        + scores.get("Address", 0) * 0.10
        + scores.get("Contact", 0) * 0.05
        + scores.get("Contact Title", 0) * 0.05
    )
    scores["Overall"] = int(weighted)
    return scores


# ─── Similar Merchant Finder ─────────────────────────────────────────────────

def find_similar_merchants(query, all_names, threshold=60):
    query_tokens = set(extract_significant_tokens(query))
    if not query_tokens:
        return []
    scored = []
    for name in all_names:
        name_tokens = set(extract_significant_tokens(name))
        if not name_tokens:
            continue
        jaccard = len(query_tokens & name_tokens) / len(query_tokens | name_tokens)
        ratio = fuzz.token_sort_ratio(query.upper(), name.upper()) / 100
        combined = jaccard * 0.5 + ratio * 0.5
        if combined * 100 >= threshold:
            scored.append((name, int(combined * 100)))
    scored.sort(key=lambda x: -x[1])
    return scored


# ─── Email Discovery ─────────────────────────────────────────────────────────

def discover_emails(records, merchant_name):
    emails = []
    for rec in records:
        if rec["merchant_name"] == merchant_name and rec["email"]:
            emails.append(rec["email"])
    unique = list(dict.fromkeys(emails))
    return unique


# ─── Duplicate Detection ─────────────────────────────────────────────────────

def find_duplicates(records):
    groups = defaultdict(list)
    for i, rec in enumerate(records):
        key = rec["merchant_name"].upper().strip()
        if key:
            groups[key].append(rec)
    return {k: v for k, v in groups.items() if len(v) > 1}


# ─── Automatic Learning ──────────────────────────────────────────────────────

class LearningSystem:
    def __init__(self, learning_file):
        self.learning_file = learning_file
        self.knowledge = {}
        self._load()

    def _load(self):
        if os.path.exists(self.learning_file):
            try:
                with open(self.learning_file, "r", encoding="utf-8") as f:
                    self.knowledge = json.load(f)
            except (json.JSONDecodeError, UnicodeDecodeError):
                self.knowledge = {}

    def save(self):
        with open(self.learning_file, "w", encoding="utf-8") as f:
            json.dump(self.knowledge, f, indent=2, ensure_ascii=False)

    def learn(self, canonical, discovered_variants):
        if canonical not in self.knowledge:
            self.knowledge[canonical] = {"aliases": [], "first_seen": datetime.now().isoformat()}
        existing = set(self.knowledge[canonical]["aliases"])
        for v in discovered_variants:
            if v not in existing:
                self.knowledge[canonical]["aliases"].append(v)
                existing.add(v)
        self.knowledge[canonical]["last_seen"] = datetime.now().isoformat()
        self.save()

    def get_aliases(self, canonical):
        entry = self.knowledge.get(canonical, {})
        return entry.get("aliases", [])

    def lookup(self, query):
        query_upper = query.upper().strip()
        for canonical, info in self.knowledge.items():
            if canonical.upper() == query_upper:
                return canonical
            for alias in info.get("aliases", []):
                if alias.upper() == query_upper:
                    return canonical
            for alias in info.get("aliases", []):
                if query_upper in alias.upper() or alias.upper() in query_upper:
                    return canonical
        return None


# ─── Report Generator ────────────────────────────────────────────────────────

class ReportGenerator:
    def __init__(self, path):
        self.path = path
        self.wb = openpyxl.Workbook()

    def _style_header(self, ws, row=1):
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_width(self, ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

    def write_sheet(self, name, headers, rows, color=None):
        if name in self.wb.sheetnames:
            del self.wb[name]
        ws = self.wb.create_sheet(name)
        ws.append(headers)
        self._style_header(ws)
        for row in rows:
            ws.append(row)
        if color and len(rows) > 0:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for cell in ws[1]:
                cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        self._auto_width(ws)
        return ws

    def save(self):
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        self.wb.save(self.path)


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ═══════════════════════════════════════════════════════════════════════════════

class MerchantIntel:
    def __init__(self, workbook_path, db_path=None, report_path=None, learning_file=None):
        self.workbook_path = workbook_path
        self.db_path = db_path or DB_PATH
        self.report_path = report_path or REPORT_PATH
        self.learning_file = learning_file or LEARNING_FILE
        self.loader = None
        self.db = None
        self.learning = None
        self.all_records = []
        self.all_merchant_names = []

    def initialize(self):
        print("=" * 70)
        print("  MERCHANT INTELLIGENCE SEARCH TOOL")
        print("=" * 70)
        print(f"\n[1/10] Discovering workbook: {self.workbook_path}")
        self.loader = WorkbookLoader(self.workbook_path)
        self.all_records = self.loader.records
        print(f"  Found {len(self.loader.sheets)} sheets, {len(self.all_records)} total records")

        print("\n[2/10] Building SQLite FTS5 database...")
        self.db = MerchantDB(self.db_path)
        self.db.clear_data()
        for rec in self.all_records:
            self.db.insert_merchant(rec)

        print("\n[3/10] Loading alias knowledge base...")
        self.learning = LearningSystem(self.learning_file)

        self.all_merchant_names = self.db.get_all_merchant_names()
        print(f"  Found {len(self.all_merchant_names)} unique merchant names")

    def run_alias_engine(self):
        print("\n[4/10] Running alias engine...")
        alias_count = 0
        for name in self.all_merchant_names:
            aliases = generate_aliases(name)
            for alias in aliases:
                self.db.insert_alias(name, alias)
                alias_count += 1
            learned = self.learning.get_aliases(name)
            for la in learned:
                self.db.insert_alias(name, la)
                alias_count += 1
        print(f"  Generated {alias_count} aliases")

    def run_learning_merge(self):
        print("\n[5/10] Merging learned aliases...")
        merged = 0
        for rec in self.all_records:
            name = rec["merchant_name"]
            if not name:
                continue
            matched = self.learning.lookup(name)
            if matched and matched.upper() != name.upper():
                discovered_variants = generate_aliases(name)
                self.learning.learn(matched, discovered_variants)
                self.db.log_learning(name, matched, 0.9)
                merged += 1
        print(f"  Merged {merged} variant names into known merchants")

    def search(self, query):
        if not query:
            return []
        query_clean = query.strip()
        results = []

        print(f"\n  Searching for: '{query_clean}'")

        print("  [Phase 3] Token intelligence...")
        tokens = extract_significant_tokens(query_clean)
        print(f"    Significant tokens: {tokens}")

        print("  [Phase 4] Full-row search across all fields...")

        scored_records = []
        for rec in self.all_records:
            scores = score_record(query_clean, rec)
            if scores["Overall"] >= 30:
                scored_records.append((scores["Overall"], scores, rec))

        scored_records.sort(key=lambda x: -x[0])

        print(f"\n  [Phase 5] Scoring results ({len(scored_records)} records scored >= 30)...")

        print("\n  [Phase 6] Discovering emails for matched merchants...")
        matched_names = set()
        for _, _, rec in scored_records[:50]:
            if rec["merchant_name"]:
                matched_names.add(rec["merchant_name"])

        email_map = {}
        for name in matched_names:
            emails = discover_emails(self.all_records, name)
            if emails:
                email_map[name] = emails
                print(f"    {name}: {', '.join(emails[:5])}{'...' if len(emails) > 5 else ''}")

        print("\n  [Phase 7] Finding similar merchants...")
        similar = find_similar_merchants(query_clean, self.all_merchant_names, threshold=55)
        for sname, sscore in similar[:10]:
            print(f"    {sname} ({sscore}%)")

        print("\n  [Phase 8] Detecting duplicates...")
        dup_groups = find_duplicates(self.all_records)
        search_tokens = set(tokens)
        relevant_dups = {}
        for dname, drecs in dup_groups.items():
            d_tokens = set(extract_significant_tokens(dname))
            if search_tokens & d_tokens or fuzz.partial_ratio(query_clean.upper(), dname.upper()) > 60:
                relevant_dups[dname] = drecs
        for dname, drecs in relevant_dups.items():
            print(f"    {dname}: {len(drecs)} occurrences")

        return {
            "query": query_clean,
            "tokens": tokens,
            "scored_records": scored_records,
            "email_map": email_map,
            "similar": similar,
            "duplicates": relevant_dups,
        }

    def generate_report(self, results):
        print(f"\n  [Phase 9] Generating Merchant Intelligence Report...")
        report = ReportGenerator(self.report_path)

        sorted_recs = results["scored_records"]

        exact = []
        high = []
        possible = []
        not_found = []
        all_emails = []
        all_phones = []
        all_contacts = []
        all_addresses = []

        seen_emails = set()
        seen_phones = set()
        seen_contacts = set()
        seen_addresses = set()

        for score_total, scores, rec in sorted_recs:
            row = [
                rec["sheet_name"], rec["row_num"], rec["merchant_name"],
                rec["slip_header"], rec["account_name"], rec["physical_addr"],
                rec["email"], rec["mobile_phone"], rec["contact_name"],
                rec["terminal_id"], rec["merchant_id"],
                scores.get("Merchant Name", 0), scores.get("Slip Header", 0),
                scores.get("Account Name", 0), scores.get("Address", 0),
                scores.get("Email", 0), scores.get("Contact", 0),
                score_total,
            ]
            if score_total == 100:
                exact.append(row)
            elif score_total >= 75:
                high.append(row)
            elif score_total >= 30:
                possible.append(row)
            else:
                not_found.append(row)

            if rec["email"] and rec["email"] not in seen_emails:
                all_emails.append([rec["merchant_name"], rec["email"], rec["sheet_name"], rec["row_num"]])
                seen_emails.add(rec["email"])
            if rec["mobile_phone"] and rec["mobile_phone"] not in seen_phones:
                all_phones.append([rec["merchant_name"], rec["mobile_phone"], rec["sheet_name"], rec["row_num"]])
                seen_phones.add(rec["mobile_phone"])
            if rec["contact_name"] and rec["contact_name"] not in seen_contacts:
                all_contacts.append([rec["merchant_name"], rec["contact_name"], rec["contact_title"], rec["sheet_name"], rec["row_num"]])
                seen_contacts.add(rec["contact_name"])
            if rec["physical_addr"] and rec["physical_addr"] not in seen_addresses:
                all_addresses.append([rec["merchant_name"], rec["physical_addr"], rec["sheet_name"], rec["row_num"]])
                seen_addresses.add(rec["physical_addr"])

        headers = ["Sheet", "Row", "Merchant Name", "Slip Header", "Account Name",
                    "Address", "Email", "Phone", "Contact", "Terminal ID", "Merchant ID",
                    "Name Score", "Slip Score", "Acct Score", "Addr Score",
                    "Email Score", "Contact Score", "Overall"]

        report.write_sheet("Exact Matches", headers, exact, color="C6EFCE") if exact else None
        report.write_sheet("High Confidence", headers, high, color="D9E2F3") if high else None
        report.write_sheet("Possible Matches", headers, possible) if possible else None
        report.write_sheet("Emails", ["Merchant Name", "Email", "Sheet", "Row"], all_emails)
        report.write_sheet("Phone Numbers", ["Merchant Name", "Phone", "Sheet", "Row"], all_phones)
        report.write_sheet("Contacts", ["Merchant Name", "Contact Name", "Title", "Sheet", "Row"], all_contacts)
        report.write_sheet("Addresses", ["Merchant Name", "Address", "Sheet", "Row"], all_addresses)

        dup_rows = []
        for dname, drecs in results["duplicates"].items():
            locations = [f"{r['sheet_name']} row {r['row_num']}" for r in drecs]
            dup_rows.append([dname, len(drecs), ", ".join(locations)])
        report.write_sheet("Duplicate Merchants", ["Merchant Name", "Count", "Locations"], dup_rows)

        report.write_sheet("Merchants Not Found", headers, not_found) if not_found else None

        summary_rows = [
            ["Total Records", str(len(sorted_recs))],
            ["Exact Matches (100%)", str(len(exact))],
            ["High Confidence (75-99)", str(len(high))],
            ["Possible Matches (30-74)", str(len(possible))],
            ["Not Found (< 30)", str(len(not_found))],
            ["Unique Emails Found", str(len(all_emails))],
            ["Unique Phones Found", str(len(all_phones))],
            ["Unique Contacts Found", str(len(all_contacts))],
            ["Unique Addresses Found", str(len(all_addresses))],
            ["Duplicate Groups", str(len(results["duplicates"]))],
            ["Similar Merchants Found", str(len(results["similar"]))],
            ["Query", results["query"]],
        ]
        report.write_sheet("Summary", ["Metric", "Value"], summary_rows)
        report.save()
        print(f"  Report saved to: {self.report_path}")

    def learn_from_results(self, results):
        print("\n  [Phase 10] Automatic learning...")
        learned_count = 0
        for _, scores, rec in results["scored_records"]:
            if scores["Overall"] >= 80 and rec["merchant_name"]:
                aliases = generate_aliases(rec["merchant_name"])
                self.learning.learn(rec["merchant_name"], aliases)
                learned_count += 1
        print(f"  Learned {learned_count} merchant patterns")
        for sim_name, sim_score in results["similar"]:
            aliases = generate_aliases(sim_name)
            self.learning.learn(sim_name, aliases)

    def interactive_mode(self):
        print("\n" + "=" * 70)
        print("  Interactive Search Mode")
        print("  Type 'exit' to quit, 'report <query>' to generate report")
        print("=" * 70)

        while True:
            try:
                query = input("\n🔍 Search: ").strip()
                if not query:
                    continue
                if query.lower() == "exit":
                    break
                if query.lower().startswith("report "):
                    q = query[7:].strip()
                    results = self.search(q)
                    self.generate_report(results)
                    self.learn_from_results(results)
                    print(f"\n  Report generated for '{q}'")
                    continue

                results = self.search(query)
                top = results["scored_records"][:15]

                print(f"\n  ┌─── Top {len(top)} Results ──────────────────────────────────────────────")
                for i, (score_total, scores, rec) in enumerate(top, 1):
                    name = rec["merchant_name"] or "(unnamed)"
                    sheet = rec["sheet_name"]
                    row = rec["row_num"]
                    print(f"  │ {i:2d}. [{score_total:3d}%] {name[:50]:50s}")
                    print(f"  │     Sheet: {sheet} row {row}")
                    if scores["Merchant Name"] > 0:
                        print(f"  │     Name={scores['Merchant Name']} Slip={scores['Slip Header']} "
                              f"Acct={scores['Account Name']} Addr={scores['Address']} "
                              f"Email={scores['Email']}")
                    if rec["email"]:
                        print(f"  │     Email: {rec['email']}")
                    if rec["mobile_phone"]:
                        print(f"  │     Phone: {rec['mobile_phone']}")
                    if rec["terminal_id"]:
                        print(f"  │     TID: {rec['terminal_id']} MID: {rec['merchant_id']}")
                    print(f"  │")

                if results["similar"]:
                    print(f"  ├─── Similar Merchants ─────────────────────────────────────────────")
                    for sname, sscore in results["similar"][:8]:
                        print(f"  │   {sname[:55]:55s} ({sscore}%)")
                    print(f"  │")

                if results["email_map"]:
                    print(f"  ├─── Emails Discovered ─────────────────────────────────────────────")
                    for mname, emails in list(results["email_map"].items())[:3]:
                        print(f"  │   {mname[:40]:40s}")
                        for e in emails[:3]:
                            print(f"  │     └─ {e}")
                    print(f"  │")

                dup_count = len(results["duplicates"])
                if dup_count:
                    print(f"  ├─── Duplicate Groups: {dup_count} ──────────────────────────────────")
                    for dname, drecs in list(results["duplicates"].items())[:3]:
                        locations = [f"{r['sheet_name']} row {r['row_num']}" for r in drecs[:5]]
                        print(f"  │   {dname[:45]:45s} ({len(drecs)}x)")
                        for loc in locations:
                            print(f"  │     └─ {loc}")
                    print(f"  │")

                print(f"  └──────────────────────────────────────────────────────────────────")

            except KeyboardInterrupt:
                print("\n  Exiting...")
                break
            except Exception as e:
                print(f"  Error: {e}")

    def close(self):
        if self.db:
            self.db.close()
        if self.learning:
            self.learning.save()


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    tool = MerchantIntel(WORKBOOK)
    try:
        tool.initialize()
        tool.run_alias_engine()
        tool.run_learning_merge()
        tool.interactive_mode()
    finally:
        tool.close()


if __name__ == "__main__":
    main()
