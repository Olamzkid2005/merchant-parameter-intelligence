"""Final MID verdict for the exported Medplus TIDs.

Findings from the tiebreaker:
  - CHANGE / SAMEDAY / NIBSS sheets AGREE with each other and their MID
    state-code matches the store's real state (LA for Lagos stores, etc.).
  - The 2ISW_Parameter sheet's merchant_id for these MEDPLUS rows is
    systematically shifted/wrong (e.g. a Lagos store gets ...7AB10 = Abia).
So the reliable MID = the value agreed by CHANGE/SAMEDAY/NIBSS.
"""
import sys, sqlite3, collections
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df = pd.read_excel('data/medplus_tids.xlsx')
rows = []
for _, r in df.iterrows():
    tid = str(r['TID']).strip()
    addr = str(r[' Address']) if ' Address' in r else str(r['Address'])
    rows.append({'store': str(r['Store Name']).strip(), 'tid': tid, 'address': addr})

GOOD_CAMPS = ('CHANGE', 'SAMEDAY', 'NIBSS')

conn = sqlite3.connect('data/intelligence.db')
c = conn.cursor()

def fam(s):
    s = s.split(' :: ')[-1].strip()
    if 'NIBSS' in s: return 'NIBSS'
    if s.startswith('2ISW_Parameter'): return 'PARAM'
    if 'Change of merchant' in s: return 'CHANGE'
    if 'Sameday' in s: return 'SAMEDAY'
    return 'OTHER'

results = []
per_tid = collections.defaultdict(list)
for t in sorted({r['tid'] for r in rows if r['tid']}):
    mids = collections.defaultdict(set)
    for mid, sheet in c.execute(
            "SELECT merchant_id, sheet_name FROM merchants WHERE tid = ? AND merchant_id != '' AND merchant_id != '0'",
            (t,)):
        mids[fam(sheet)].add(mid)
    good = set()
    for camp in GOOD_CAMPS:
        good |= mids.get(camp, set())
    good.discard('0')
    # exclude the KN010 garbage seen in CHANGE (YISW123456KN010 is a typo variant)
    good = {m for m in good if not m.startswith('Y')}
    param = mids.get('PARAM', set())
    verdict = {
        'tid': t,
        'good_mids': sorted(good),
        'param_mids': sorted(param),
        'agreed': len(good) == 1,
    }
    per_tid[t] = verdict

# Build final table
final = []
agreed_cnt = 0
multi_cnt = 0
none_cnt = 0
for r in rows:
    tid = r['tid']
    v = per_tid.get(tid)
    if not v or not v['good_mids']:
        mid = ''
        status = 'NO RELIABLE MID'
        none_cnt += 1
    elif v['agreed']:
        mid = v['good_mids'][0]
        status = 'OK'
        agreed_cnt += 1
    else:
        mid = ', '.join(v['good_mids'])
        status = 'MULTIPLE'
        multi_cnt += 1
    final.append({'Store': r['store'], 'TID': tid, 'MID': mid, 'Status': status,
                  'Address': r['address']})

print(f"total rows: {len(final)}")
print(f"  OK (single agreed MID):     {agreed_cnt}")
print(f"  MULTIPLE (still clash):     {multi_cnt}")
print(f"  NO RELIABLE MID:            {none_cnt}")

if multi_cnt:
    print("\n=== MULTIPLE (even among good sheets) ===")
    seen = set()
    for f in final:
        if f['Status'] == 'MULTIPLE' and f['TID'] not in seen:
            seen.add(f['TID'])
            print(f"  {f['TID']}: {f['MID']}  ({f['Store']})")

# Save
wb = Workbook()
ws = wb.active
ws.title = 'Medplus MIDs'
headers = ['Store Name', 'TID', 'MID (Merchant ID)', 'Status', 'Address']
ws.append(headers)
hdr_fill = PatternFill('solid', fgColor='1F4E79')
thin = Side(style='thin', color='D9D9D9')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = hdr_fill
    cell.font = Font(bold=True, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center')
    cell.border = border
for f in final:
    ws.append([f['Store'], f['TID'], f['MID'], f['Status'], f['Address']])
    rr = ws.max_row
    for col in range(1, len(headers) + 1):
        ws.cell(row=rr, column=col).border = border
for i, w in enumerate([30, 11, 20, 16, 60], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:E{ws.max_row}'
out = 'data/medplus_mids.xlsx'
wb.save(out)
print(f"\nsaved: {out}")

conn.close()
