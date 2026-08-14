import pandas as pd
import openpyxl

file = r"C:\Users\David.Olamijulo\downloads\parameter\data\2ISW_Parameter_File 5.xlsx"

wb = openpyxl.load_workbook(file)
print("Sheet names:", wb.sheetnames)
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== Sheet: {sheet_name} ===")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    print()
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(f"Row {i}: {row}")
        if i >= 30:
            break
    print("\n")
