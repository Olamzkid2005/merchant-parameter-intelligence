from openpyxl import load_workbook

file = r"C:\Users\David.Olamijulo\downloads\parameter\data\2ISW_Parameter_File 5.xlsx"

wb = load_workbook(file, read_only=True)
print("Sheet names:", wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== {name} ===")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    # First 5 rows
    count = 0
    for row in ws.iter_rows(values_only=True):
        print(f"  Row {count+1}: {row}")
        count += 1
        if count >= 5:
            break
    # Count total rows quickly
    total = sum(1 for _ in ws.iter_rows(values_only=True))
    print(f"Total data rows: {total - 1} (excluding header)")

wb.close()
